import csv
import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO

from django.contrib.auth.models import Group, User
from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Sum
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.test import APITestCase

from .models import (
    ActividadTrabajo,
    Almacen,
    Cliente,
    Compra,
    CompraDetalle,
    DetalleDocumentoEstandarizado,
    DetalleSupervisor,
    MedidaCorrectiva,
    ConexionDetalleDocumento,
    Dimension,
    GestionCambio,
    HistorialConsumible,
    HistorialUbicacionItem,
    IPERC,
    Item,
    ItemUnidad,
    LoteConsumible,
    Maquinaria,
    MovimientoConsumible,
    MovimientoRepuesto,
    EncabezadoDocumentoEstandarizacion,
    OrdenRequerimiento,
    OrdenRequerimientoDetalle,
    OrdenTrabajo,
    PerfilUsuario,
    Proveedor,
    ReporteOrden,
    ReporteIPERC,
    SecuenciaControlRiesgo,
    Sistema,
    ActividadChecklist,
    Checklist,
    ChecklistActividad,
    ChecklistEjecucion,
    ChecklistRespuesta,
    TareaPorEstandarizar,
    Trabajador,
    UbicacionCliente,
    UnidadMedida,
    current_local_date,
)
from .serializers import OrdenTrabajoSerializer


class TimezoneConfigurationTests(APITestCase):
    def test_backend_usa_zona_horaria_de_peru_y_fecha_local_en_ot(self):
        self.assertEqual(settings.TIME_ZONE, "America/Lima")
        self.assertFalse(settings.USE_TZ)
        self.assertIs(OrdenTrabajo._meta.get_field("fecha").default, current_local_date)


class EstandarizacionModelsTests(APITestCase):
    def test_encabezado_documento_autocompleta_codigo_area_y_revision(self):
        user = User.objects.create_user(username="estandar-user", password="secret123")
        trabajador = Trabajador.objects.create(
            nombres="Lucia",
            apellidos="Ramos",
            dni="55667788",
            puesto="Supervisora",
        )
        PerfilUsuario.objects.create(user=user, trabajador=trabajador)

        tarea = TareaPorEstandarizar.objects.create(
            codigo="TES-001",
            nombre_tarea="Cambio de filtro principal",
            nivel_criticidad=TareaPorEstandarizar.NivelCriticidad.ALTO,
            desarrollado=True,
            area="Taller",
        )

        encabezado = EncabezadoDocumentoEstandarizacion.objects.create(
            tarea_por_estandarizar=tarea,
            creado_por=user,
            fecha=date(2026, 5, 20),
        )

        self.assertTrue(encabezado.codigo.startswith("DE-2026-"))
        self.assertEqual(encabezado.area, "Taller")
        self.assertEqual(encabezado.revision_id, trabajador.id)


class TareaPorEstandarizarViewSetTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="jefe-procesos",
            password="secret123",
        )
        group, _ = Group.objects.get_or_create(name="Jefe de Tecnicos")
        self.user.groups.add(group)
        self.client.force_authenticate(user=self.user)

        self.dimension = Dimension.objects.create(
            codigo="UNI",
            nombre="Unidad",
            descripcion="Conteo de items",
            activo=True,
        )
        self.unidad = UnidadMedida.objects.create(
            nombre="CANTIDAD",
            simbolo="und",
            dimension=self.dimension,
            es_base=True,
            activo=True,
        )
        self.item = Item.objects.create(
            codigo="REP-001",
            nombre="Filtro de aceite",
            tipo_insumo=Item.TipoInsumo.REPUESTO,
            dimension=self.dimension,
            unidad_medida=self.unidad,
        )

    def test_jefe_tecnicos_puede_listar_y_crear_tareas_por_estandarizar(self):
        TareaPorEstandarizar.objects.create(
            codigo="PROC-001",
            nombre_tarea="Inspeccion inicial",
            nivel_criticidad=TareaPorEstandarizar.NivelCriticidad.ALTO,
            desarrollado=False,
            area="Taller",
            item=self.item,
        )

        response = self.client.get("/api/tareas-por-estandarizar/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["item_codigo"], "REP-001")

        create_response = self.client.post(
            "/api/tareas-por-estandarizar/",
            {
                "codigo": "proc-002",
                "nombre_tarea": "Cambio de filtro",
                "nivel_criticidad": TareaPorEstandarizar.NivelCriticidad.MEDIO,
                "desarrollado": True,
                "area": "Campo",
                "item": self.item.id,
            },
            format="json",
        )
        self.assertEqual(create_response.status_code, 201)
        self.assertEqual(create_response.data["codigo"], "PROC-002")
        self.assertEqual(create_response.data["desarrollado_label"], "Si")

    def test_documento_estandarizacion_usa_codigo_de_la_tarea_y_lista_detalles(self):
        tarea = TareaPorEstandarizar.objects.create(
            codigo="PROC-010",
            nombre_tarea="Estandarizar cambio de aceite",
            nivel_criticidad=TareaPorEstandarizar.NivelCriticidad.ALTO,
            desarrollado=False,
            area="Taller",
            item=self.item,
        )

        encabezado_response = self.client.post(
            "/api/encabezados-estandarizacion/",
            {
                "tarea_por_estandarizar": tarea.id,
                "fecha": "2026-05-22",
            },
            format="json",
        )
        self.assertEqual(encabezado_response.status_code, 201)
        self.assertEqual(encabezado_response.data["codigo"], "PROC-010")
        self.assertEqual(encabezado_response.data["area"], "Taller")

        detalle_response = self.client.post(
            "/api/detalles-estandarizacion/",
            {
                "encabezado_documento": encabezado_response.data["id"],
                "numero": 1,
                "recurso": "Llave de filtro",
                "actividad": "Retirar filtro usado",
                "detalle_actividad": "Aflojar el filtro con la herramienta adecuada.",
                "responsable": "Tecnico lider",
                "nota_importante": "Verificar bandeja de goteo.",
                "consideraciones": "Usar guantes resistentes al aceite.",
            },
            format="json",
        )
        self.assertEqual(detalle_response.status_code, 201)

        list_response = self.client.get(
            f"/api/detalles-estandarizacion/?encabezado_documento={encabezado_response.data['id']}"
        )
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(len(list_response.data), 1)
        self.assertEqual(list_response.data[0]["actividad"], "Retirar filtro usado")

    def test_vista_flujo_crea_bloques_minimos_y_conexion_automatica(self):
        tarea = TareaPorEstandarizar.objects.create(
            codigo="PROC-011",
            nombre_tarea="Flujo vertical",
            nivel_criticidad=TareaPorEstandarizar.NivelCriticidad.MEDIO,
            desarrollado=False,
            area="Taller",
        )
        encabezado = EncabezadoDocumentoEstandarizacion.objects.create(
            tarea_por_estandarizar=tarea,
            creado_por=self.user,
            fecha=date(2026, 5, 22),
        )

        first_response = self.client.post(
            "/api/detalles-estandarizacion/crear_desde_flujo/",
            {
                "encabezado_documento": encabezado.id,
                "actividad": "Inicio del procedimiento",
                "tipo_nodo": "inicio",
            },
            format="json",
        )
        self.assertEqual(first_response.status_code, 201)
        self.assertEqual(first_response.data["numero"], 1)
        self.assertEqual(first_response.data["tipo_nodo"], "inicio")
        self.assertEqual(first_response.data["recurso"], "")

        second_response = self.client.post(
            "/api/detalles-estandarizacion/crear_desde_flujo/",
            {
                "encabezado_documento": encabezado.id,
                "actividad": "Evaluar presion del sistema",
                "tipo_nodo": "decision",
            },
            format="json",
        )
        self.assertEqual(second_response.status_code, 201)
        self.assertEqual(second_response.data["numero"], 2)

        conexiones = ConexionDetalleDocumento.objects.filter(documento=encabezado)
        self.assertEqual(conexiones.count(), 1)
        conexion = conexiones.first()
        self.assertEqual(conexion.origen.actividad, "Inicio del procedimiento")
        self.assertEqual(conexion.destino.actividad, "Evaluar presion del sistema")

    def test_eliminar_nodo_renumera_la_secuencia_del_flujo(self):
        tarea = TareaPorEstandarizar.objects.create(
            codigo="PROC-012",
            nombre_tarea="Secuencia de bloques",
            nivel_criticidad=TareaPorEstandarizar.NivelCriticidad.BAJO,
            desarrollado=False,
            area="Campo",
        )
        encabezado = EncabezadoDocumentoEstandarizacion.objects.create(
            tarea_por_estandarizar=tarea,
            creado_por=self.user,
            fecha=date(2026, 5, 22),
        )
        detalle_1 = DetalleDocumentoEstandarizado.objects.create(
            encabezado_documento=encabezado,
            numero=1,
            actividad="Nodo 1",
            posicion_y=0,
        )
        detalle_2 = DetalleDocumentoEstandarizado.objects.create(
            encabezado_documento=encabezado,
            numero=2,
            actividad="Nodo 2",
            posicion_y=180,
        )
        detalle_3 = DetalleDocumentoEstandarizado.objects.create(
            encabezado_documento=encabezado,
            numero=3,
            actividad="Nodo 3",
            posicion_y=360,
        )
        ConexionDetalleDocumento.objects.create(
            documento=encabezado,
            origen=detalle_1,
            destino=detalle_2,
        )
        ConexionDetalleDocumento.objects.create(
            documento=encabezado,
            origen=detalle_2,
            destino=detalle_3,
        )

        delete_response = self.client.delete(
            f"/api/detalles-estandarizacion/{detalle_2.id}/"
        )
        self.assertEqual(delete_response.status_code, 204)

        remaining = list(
            DetalleDocumentoEstandarizado.objects
            .filter(encabezado_documento=encabezado)
            .order_by("numero", "id")
            .values_list("actividad", "numero", "posicion_y")
        )
        self.assertEqual(
            remaining,
            [
                ("Nodo 1", 1, 0.0),
                ("Nodo 3", 2, 180.0),
            ],
        )


class ReporteOrdenViewSetTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="jefe-reportes",
            password="secret123",
        )
        group, _ = Group.objects.get_or_create(name="Jefe de Tecnicos")
        self.user.groups.add(group)
        self.client.force_authenticate(user=self.user)

        self.maquinaria = Maquinaria.objects.create(
            codigo_maquina="MQ-REPORT-01",
            nombre="Excavadora de campo",
            descripcion="Unidad de pruebas",
        )

    def test_crea_reporte_automatico_al_registrar_ot_en_campo(self):
        response = self.client.post(
            "/api/trabajos/",
            {
                "maquinaria": self.maquinaria.id,
                "prioridad": "URGENTE",
                "lugar": "CAMPO",
                "ubicacion_detalle": "Frente 4",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        orden = OrdenTrabajo.objects.get(pk=response.data["id"])
        reporte = ReporteOrden.objects.filter(orden_trabajo=orden).first()
        self.assertIsNotNone(reporte)
        self.assertEqual(reporte.estado, ReporteOrden.Estado.PENDIENTE)
        self.assertTrue(reporte.codigo.startswith("RTO-"))
        self.assertEqual(reporte.detalles_supervisor.count(), 3)
        self.assertEqual(
            reporte.detalles_supervisor.filter(tipo=DetalleSupervisor.Tipo.AUTORIZA).count(),
            1,
        )
        self.assertEqual(
            reporte.detalles_supervisor.filter(tipo=DetalleSupervisor.Tipo.VERIFICA).count(),
            1,
        )
        self.assertEqual(reporte.detalles_supervisor.filter(tipo="").count(), 1)

    def test_no_crea_reporte_automatico_si_ot_es_en_taller(self):
        response = self.client.post(
            "/api/trabajos/",
            {
                "maquinaria": self.maquinaria.id,
                "prioridad": "REGULAR",
                "lugar": "TALLER",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        orden = OrdenTrabajo.objects.get(pk=response.data["id"])
        self.assertFalse(ReporteOrden.objects.filter(orden_trabajo=orden).exists())

    def test_lista_y_crea_reportes_solo_para_ot_de_campo(self):
        orden_campo = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            prioridad="URGENTE",
            lugar=OrdenTrabajo.Lugar.CAMPO,
            ubicacion_detalle="Cantera norte",
        )

        orden_taller = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            prioridad="REGULAR",
            lugar=OrdenTrabajo.Lugar.TALLER,
        )

        list_response = self.client.get("/api/reportes-orden/")
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(len(list_response.data), 1)
        self.assertEqual(list_response.data[0]["orden_trabajo_codigo"], orden_campo.codigo_orden)

        invalid_create = self.client.post(
            "/api/reportes-orden/",
            {
                "orden_trabajo": orden_taller.id,
                "fecha": str(orden_taller.fecha),
            },
            format="json",
        )
        self.assertEqual(invalid_create.status_code, 400)
        self.assertIn("orden_trabajo", invalid_create.data)

        valid_create = self.client.post(
            "/api/reportes-orden/",
            {
                "orden_trabajo": orden_campo.id,
                "fecha": str(orden_campo.fecha),
                "epp": "Casco, guantes",
            },
            format="json",
        )
        self.assertEqual(valid_create.status_code, 400)

        otro_campo = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            prioridad="EMERGENCIA",
            lugar=OrdenTrabajo.Lugar.CAMPO,
            ubicacion_detalle="Acceso sur",
        )
        self.assertTrue(ReporteOrden.objects.filter(orden_trabajo=otro_campo).exists())

        second_list_response = self.client.get("/api/reportes-orden/")
        self.assertEqual(second_list_response.status_code, 200)
        self.assertEqual(len(second_list_response.data), 2)

    def test_actualizar_supervisor_verifica_regenera_slot_vacio(self):
        orden_campo = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            prioridad="URGENTE",
            lugar=OrdenTrabajo.Lugar.CAMPO,
            ubicacion_detalle="Cantera norte",
        )
        reporte = ReporteOrden.objects.get(orden_trabajo=orden_campo)
        placeholder = reporte.detalles_supervisor.filter(tipo="").first()

        response = self.client.patch(
            f"/api/detalles-supervisor/{placeholder.id}/",
            {
                "tipo": DetalleSupervisor.Tipo.VERIFICA,
                "nombres": "Luis",
                "apellidos": "Rojas",
                "dni": "99887766",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        reporte.refresh_from_db()
        self.assertGreaterEqual(
            reporte.detalles_supervisor.filter(tipo=DetalleSupervisor.Tipo.VERIFICA).count(),
            2,
        )
        self.assertEqual(reporte.detalles_supervisor.filter(tipo="").count(), 1)

    def test_crear_iperc_desde_reporte_genera_registro_con_motivo_por_defecto(self):
        orden_campo = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            prioridad="URGENTE",
            lugar=OrdenTrabajo.Lugar.CAMPO,
            ubicacion_detalle="Cantera norte",
        )
        reporte = ReporteOrden.objects.get(orden_trabajo=orden_campo)

        response = self.client.post(
            f"/api/reportes-orden/{reporte.id}/crear_iperc/",
            {},
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        reporte_iperc = ReporteIPERC.objects.get(pk=response.data["id"])
        self.assertEqual(reporte_iperc.orden_trabajo_id, orden_campo.id)
        self.assertEqual(reporte_iperc.motivo, ReporteIPERC.Motivo.ACTIVIDAD_NUEVA)
        self.assertEqual(
            reporte_iperc.tarea,
            f"{orden_campo.codigo_orden} - {self.maquinaria.nombre}",
        )
        self.assertEqual(reporte_iperc.ipercs.count(), 1)
        placeholder = reporte.detalles_supervisor.filter(tipo="").first()
        self.assertIsNotNone(placeholder)
        placeholder.refresh_from_db()
        self.assertEqual(placeholder.reporte_iperc_id, reporte_iperc.id)
        self.assertEqual(response.data["motivo_label"], "Actividad nueva")
        self.assertEqual(
            response.data["tarea"],
            f"{orden_campo.codigo_orden} - {self.maquinaria.nombre}",
        )
        self.assertEqual(response.data["ipercs_count"], 1)
        self.assertIsNotNone(response.data["supervisor_pendiente"])
        self.assertEqual(
            response.data["supervisor_pendiente"]["reporte_iperc"],
            reporte_iperc.id,
        )


class ChecklistViewSetTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="jefe-checklist",
            password="secret123",
        )
        group, _ = Group.objects.get_or_create(name="Jefe de Tecnicos")
        self.user.groups.add(group)
        self.client.force_authenticate(user=self.user)

        self.dimension = Dimension.objects.create(
            codigo="CHK",
            nombre="Checklist",
            descripcion="Dimension base checklist",
            activo=True,
        )
        self.unidad = UnidadMedida.objects.create(
            nombre="UNIDAD",
            simbolo="und",
            dimension=self.dimension,
            es_base=True,
            activo=True,
        )
        self.item = Item.objects.create(
            codigo="CHK-001",
            nombre="Manometro",
            tipo_insumo=Item.TipoInsumo.CONSUMIBLE,
            dimension=self.dimension,
            unidad_medida=self.unidad,
        )
        self.maquinaria = Maquinaria.objects.create(
            codigo_maquina="MQ-CHK-01",
            nombre="Compresor principal",
            descripcion="Unidad de checklist",
        )
        self.cliente = Cliente.objects.create(
            nombre="Cliente Checklist",
            ruc="20123456789",
        )
        self.ubicacion = UbicacionCliente.objects.create(
            cliente=self.cliente,
            nombre="Patio norte",
            direccion="Sector A",
        )

    def test_crea_checklist_con_actividad_y_sistema_dinamicos(self):
        response = self.client.post(
            "/api/checklists/",
            {
                "motivo": "Checklist de arranque de motor",
                "fecha": "2026-05-25",
                "estado": "ACTIVO",
                "actividades_payload": [
                    {
                        "orden": 1,
                        "obligatorio": True,
                        "actividad_data": {
                            "descripcion": "Verificar presion inicial",
                            "tipo_respuesta": "BOOLEANO",
                            "obligatorio": True,
                            "requiere_observacion": True,
                            "requiere_evidencia": False,
                            "item": self.item.id,
                            "activo": True,
                        },
                        "sistema_data": {
                            "nombre": "Motor principal",
                            "descripcion": "Sistema de propulsion",
                        },
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        checklist = Checklist.objects.get(pk=response.data["id"])
        self.assertEqual(checklist.creado_por_id, self.user.id)
        self.assertEqual(checklist.actividades_checklist.count(), 1)
        self.assertEqual(response.data["actividades_count"], 1)

        relacion = checklist.actividades_checklist.select_related("actividad", "sistema").first()
        self.assertIsNotNone(relacion)
        self.assertEqual(relacion.orden, 1)
        self.assertEqual(relacion.actividad.descripcion, "Verificar presion inicial")
        self.assertEqual(relacion.actividad.item_id, self.item.id)
        self.assertEqual(relacion.sistema.nombre, "Motor principal")

    def test_crea_checklist_sin_relaciones_iniciales(self):
        response = self.client.post(
            "/api/checklists/",
            {
                "motivo": "Checklist desacoplado para cabina",
                "fecha": "2026-05-26",
                "estado": "BORRADOR",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        checklist = Checklist.objects.get(pk=response.data["id"])
        self.assertEqual(checklist.motivo, "Checklist desacoplado para cabina")
        self.assertEqual(checklist.actividades_checklist.count(), 0)
        self.assertEqual(response.data["actividades_count"], 0)
        self.assertEqual(checklist.creado_por_id, self.user.id)

    def test_lista_ejecuciones_de_checklist_con_respuestas_count(self):
        checklist = Checklist.objects.create(
            motivo="Checklist de tablero electrico",
            fecha=date(2026, 5, 25),
            estado=Checklist.Estado.ACTIVO,
            creado_por=self.user,
        )
        actividad = ActividadChecklist.objects.create(
            descripcion="Confirmar luces de alarma",
            tipo_respuesta=ActividadChecklist.TipoRespuesta.BOOLEANO,
            obligatorio=True,
            activo=True,
        )
        sistema = Sistema.objects.create(
            nombre="Tablero electrico",
            descripcion="Controles de cabina",
        )
        checklist_actividad = ChecklistActividad.objects.create(
            checklist=checklist,
            actividad=actividad,
            sistema=sistema,
            orden=1,
            obligatorio=True,
        )
        ejecucion = ChecklistEjecucion.objects.create(
            checklist=checklist,
            fecha=date(2026, 5, 25),
            fecha_inicio=date(2026, 5, 25),
            fecha_fin=date(2026, 5, 25),
            horometro=Decimal("1450.50"),
            maquinaria=self.maquinaria,
            lugar=self.ubicacion,
            motivo="Rutina diaria",
            realizado_por=self.user,
            estado=ChecklistEjecucion.Estado.COMPLETADO,
        )
        ChecklistRespuesta.objects.create(
            ejecucion=ejecucion,
            checklist_actividad=checklist_actividad,
            vb=True,
            valor_booleano=True,
            observacion="Todo conforme",
        )

        response = self.client.get("/api/checklist-ejecuciones/")
        respuestas_response = self.client.get("/api/checklist-respuestas/")

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(len(response.data), 1)
        self.assertTrue(response.data[0]["codigo"].startswith("CHK-EJ-2026-"))
        self.assertEqual(response.data[0]["checklist_motivo"], "Checklist de tablero electrico")
        self.assertEqual(response.data[0]["realizado_por_username"], self.user.username)
        self.assertEqual(response.data[0]["estado"], ChecklistEjecucion.Estado.COMPLETADO)
        self.assertEqual(response.data[0]["respuestas_count"], 1)
        self.assertEqual(response.data[0]["maquinaria"], self.maquinaria.id)
        self.assertEqual(response.data[0]["lugar"], self.ubicacion.id)
        self.assertEqual(response.data[0]["horometro"], "1450.50")
        self.assertEqual(response.data[0]["fecha_inicio"], "2026-05-25")
        self.assertEqual(response.data[0]["fecha_fin"], "2026-05-25")
        self.assertEqual(response.data[0]["maquinaria_nombre"], "MQ-CHK-01 - Compresor principal")
        self.assertEqual(response.data[0]["lugar_nombre"], "Patio norte")
        self.assertEqual(respuestas_response.status_code, 200, respuestas_response.data)
        self.assertEqual(len(respuestas_response.data), 1)
        self.assertTrue(respuestas_response.data[0]["vb"])


class ReporteIPERCViewSetTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="jefe-iperc",
            password="secret123",
        )
        group, _ = Group.objects.get_or_create(name="Jefe de Tecnicos")
        self.user.groups.add(group)
        self.client.force_authenticate(user=self.user)

        self.maquinaria = Maquinaria.objects.create(
            codigo_maquina="MQ-IPERC-01",
            nombre="Excavadora IPERC",
            descripcion="Unidad de prueba",
        )
        self.orden = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            fecha=date(2026, 5, 22),
            prioridad="URGENTE",
            lugar=OrdenTrabajo.Lugar.CAMPO,
            ubicacion_detalle="Frente norte",
        )

    def test_crea_y_lista_reportes_iperc_con_codigo_y_fecha(self):
        response = self.client.post(
            "/api/reportes-iperc/",
            {
                "orden_trabajo": self.orden.id,
                "motivo": ReporteIPERC.Motivo.CAMBIO_PROCESO,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        reporte = ReporteIPERC.objects.get(pk=response.data["id"])
        self.assertTrue(reporte.codigo.startswith("RIPERC-2026-"))
        self.assertEqual(reporte.fecha, self.orden.fecha)
        self.assertEqual(
            reporte.tarea,
            f"{self.orden.codigo_orden} - {self.maquinaria.nombre}",
        )
        self.assertEqual(reporte.ipercs.count(), 1)

        list_response = self.client.get("/api/reportes-iperc/")
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(len(list_response.data), 1)
        self.assertEqual(list_response.data[0]["codigo"], reporte.codigo)
        self.assertEqual(list_response.data[0]["orden_trabajo_codigo"], self.orden.codigo_orden)
        self.assertEqual(list_response.data[0]["motivo"], ReporteIPERC.Motivo.CAMBIO_PROCESO)
        self.assertEqual(list_response.data[0]["motivo_label"], "Cambio del proceso")
        self.assertEqual(
            list_response.data[0]["tarea"],
            f"{self.orden.codigo_orden} - {self.maquinaria.nombre}",
        )
        self.assertEqual(list_response.data[0]["ipercs_count"], 1)

    def test_crea_reporte_iperc_manual_con_tarea_y_sin_orden(self):
        response = self.client.post(
            "/api/reportes-iperc/",
            {
                "motivo": ReporteIPERC.Motivo.TRABAJO_NO_RUTINARIO,
                "tarea": "Inspeccion extraordinaria de sistema hidraulico",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        reporte = ReporteIPERC.objects.get(pk=response.data["id"])
        self.assertIsNone(reporte.orden_trabajo)
        self.assertEqual(
            reporte.tarea,
            "Inspeccion extraordinaria de sistema hidraulico",
        )
        self.assertEqual(reporte.motivo, ReporteIPERC.Motivo.TRABAJO_NO_RUTINARIO)
        self.assertEqual(reporte.ipercs.count(), 1)

    def test_reporte_iperc_manual_exige_tarea_cuando_no_tiene_orden(self):
        response = self.client.post(
            "/api/reportes-iperc/",
            {
                "motivo": ReporteIPERC.Motivo.ACTIVIDAD_NUEVA,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertIn("tarea", response.data)

    def test_retrieve_incluye_supervisor_pendiente_y_registros_relacionados(self):
        reporte = ReporteIPERC.objects.create(
            orden_trabajo=self.orden,
            motivo=ReporteIPERC.Motivo.ACTIVIDAD_NUEVA,
        )
        supervisor = reporte.detalles_supervisor.filter(tipo="").first()
        self.assertIsNotNone(supervisor)

        iperc = reporte.ipercs.first()
        iperc.descripcion_peligro = "Aceite caliente"
        iperc.consecuencia_peligro = "Quemaduras"
        iperc.evaluacion_iperc = "ALTO"
        iperc.evaluacion_riesgo_residual = "MEDIO"
        iperc.save()

        GestionCambio.objects.create(
            iperc=iperc,
            implementacion="Aislar la zona y usar guantes termicos",
            estado=GestionCambio.Estado.APROBADO,
            observacion="Control previo al inicio",
        )
        SecuenciaControlRiesgo.objects.create(
            reporte_iperc=reporte,
            actividad="Verificar enfriamiento del sistema",
        )
        MedidaCorrectiva.objects.create(
            reporte_iperc=reporte,
            supervisor=supervisor,
            detalle="Detener la tarea ante fuga no controlada",
        )

        response = self.client.get(f"/api/reportes-iperc/{reporte.id}/")

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(
            response.data["tarea"],
            f"{self.orden.codigo_orden} - {self.maquinaria.nombre}",
        )
        self.assertIsNotNone(response.data["supervisor_pendiente"])
        self.assertEqual(response.data["supervisor_pendiente"]["id"], supervisor.id)
        self.assertEqual(len(response.data["ipercs"]), 1)
        self.assertEqual(response.data["ipercs"][0]["descripcion_peligro"], "Aceite caliente")
        self.assertEqual(response.data["ipercs"][0]["evaluacion_iperc"], "ALTO")
        self.assertEqual(response.data["ipercs"][0]["evaluacion_riesgo_residual"], "MEDIO")
        self.assertEqual(len(response.data["ipercs"][0]["gestiones_cambio"]), 1)
        self.assertEqual(
            response.data["ipercs"][0]["gestiones_cambio"][0]["implementacion"],
            "Aislar la zona y usar guantes termicos",
        )
        self.assertEqual(len(response.data["secuencias_control_riesgo"]), 1)
        self.assertEqual(
            response.data["secuencias_control_riesgo"][0]["actividad"],
            "Verificar enfriamiento del sistema",
        )
        self.assertEqual(len(response.data["medidas_correctivas"]), 1)
        self.assertEqual(
            response.data["medidas_correctivas"][0]["detalle"],
            "Detener la tarea ante fuga no controlada",
        )
        self.assertEqual(
            response.data["medidas_correctivas"][0]["supervisor"],
            supervisor.id,
        )

    def test_reporte_iperc_asociado_a_orden_conserva_tarea_derivada(self):
        reporte = ReporteIPERC.objects.create(
            orden_trabajo=self.orden,
            motivo=ReporteIPERC.Motivo.ACTIVIDAD_NUEVA,
        )

        response = self.client.patch(
            f"/api/reportes-iperc/{reporte.id}/",
            {
                "tarea": "Texto manual que no debe persistir",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200, response.data)
        reporte.refresh_from_db()
        self.assertEqual(
            reporte.tarea,
            f"{self.orden.codigo_orden} - {self.maquinaria.nombre}",
        )

    def test_iperc_registro_rechaza_evaluaciones_fuera_de_alto_medio_bajo(self):
        reporte = ReporteIPERC.objects.create(
            orden_trabajo=self.orden,
            motivo=ReporteIPERC.Motivo.ACTIVIDAD_NUEVA,
        )
        iperc = reporte.ipercs.first()

        response = self.client.patch(
            f"/api/iperc-registros/{iperc.id}/",
            {
                "reporte_iperc": reporte.id,
                "descripcion_peligro": "Faja expuesta",
                "consecuencia_peligro": "Atrapamiento",
                "evaluacion_iperc": "CRITICO",
                "evaluacion_riesgo_residual": "BAJO",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertIn("evaluacion_iperc", response.data)

    def test_reporte_iperc_incrementa_correlativo_por_anio(self):
        primero = ReporteIPERC.objects.create(
            orden_trabajo=self.orden,
            motivo=ReporteIPERC.Motivo.ACTIVIDAD_NUEVA,
        )
        segundo = ReporteIPERC.objects.create(
            orden_trabajo=self.orden,
            motivo=ReporteIPERC.Motivo.FRECUENCIA_PERIODICA,
        )

        self.assertTrue(primero.codigo.endswith("00001"))
        self.assertTrue(segundo.codigo.endswith("00002"))


class CatalogoSyncViewTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="admin-test",
            password="secret123",
        )
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        self.client.force_authenticate(user=self.user)

        self.dimension = Dimension.objects.create(
            id=1,
            codigo="LONG",
            nombre="Longitud",
            descripcion="Base",
            activo=True,
        )
        self.unidad = UnidadMedida.objects.create(
            id=1,
            nombre="METRO",
            simbolo="m",
            dimension=self.dimension,
            es_base=True,
            activo=True,
        )
        self.item = Item.objects.create(
            id=1,
            codigo="ITM-001",
            nombre="Manguera",
            tipo_insumo=Item.TipoInsumo.CONSUMIBLE,
            dimension=self.dimension,
            unidad_medida=self.unidad,
            favorito=False,
            volvo=False,
            ultimo_correlativo=3,
        )
        self.maquinaria = Maquinaria.objects.create(
            id=1,
            codigo_maquina="MQ-01",
            nombre="Excavadora",
            descripcion="Original",
            observacion="Inicial",
            gasto="12.50",
        )
        self.cliente = Cliente.objects.create(
            id=1,
            nombre="Cliente Uno",
            ruc="20111111111",
        )
        self.proveedor = Proveedor.objects.create(
            id=1,
            nombre="Proveedor Uno",
            ruc="20999999999",
            direccion="Av. Base",
        )

    def test_metadata_returns_supported_tables(self):
        response = self.client.get("/api/catalogo-sync/?metadata=1")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        keys = [table["key"] for table in payload["tables"]]

        self.assertIn("maquinarias", keys)
        self.assertIn("items", keys)
        self.assertIn("csv", payload["formats"])
        self.assertIn("xlsx", payload["formats"])

    def test_export_returns_configured_tables_with_primary_keys(self):
        response = self.client.get("/api/catalogo-sync/")

        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertEqual(payload["meta"]["record_counts"]["maquinarias"], 1)
        self.assertEqual(payload["tables"]["maquinarias"][0]["id"], self.maquinaria.id)
        self.assertEqual(payload["tables"]["clientes"][0]["id"], self.cliente.id)
        self.assertEqual(payload["tables"]["items"][0]["dimension"], self.dimension.id)
        self.assertEqual(payload["tables"]["items"][0]["unidad_medida"], self.unidad.id)

    def test_export_specific_table_in_csv_and_xlsx(self):
        csv_response = self.client.get(
            "/api/catalogo-sync/",
            {"table": "maquinarias", "file_format": "csv"},
        )

        self.assertEqual(csv_response.status_code, 200)
        csv_text = csv_response.content.decode("utf-8-sig")
        csv_rows = list(csv.DictReader(StringIO(csv_text)))
        self.assertEqual(csv_rows[0]["codigo_maquina"], "MQ-01")
        self.assertEqual(csv_rows[0]["nombre"], "Excavadora")

        xlsx_response = self.client.get(
            "/api/catalogo-sync/",
            {"table": "clientes", "file_format": "xlsx"},
        )

        self.assertEqual(xlsx_response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(xlsx_response.content))
        worksheet = workbook.active

        self.assertEqual(worksheet["A1"].value, "id")
        self.assertEqual(worksheet["B1"].value, "nombre")
        self.assertEqual(worksheet["B2"].value, "Cliente Uno")

    def test_import_upserts_records_by_primary_key(self):
        payload = {
            "tables": {
                "maquinarias": [
                    {
                        "id": 1,
                        "codigo_maquina": "MQ-01",
                        "nombre": "Excavadora 320",
                        "descripcion": "Actualizada",
                        "observacion": "Operativa",
                        "gasto": "22.50",
                    },
                    {
                        "id": 2,
                        "codigo_maquina": "MQ-02",
                        "nombre": "Motoniveladora",
                        "descripcion": "",
                        "observacion": "",
                        "gasto": "0.00",
                    },
                ],
                "clientes": [
                    {
                        "id": 1,
                        "nombre": "Cliente Uno",
                        "ruc": "20111111111",
                    },
                    {
                        "id": 2,
                        "nombre": "Cliente Dos",
                        "ruc": "20222222222",
                    },
                ],
                "dimensiones": [
                    {
                        "id": 1,
                        "codigo": "LONG",
                        "nombre": "Longitud",
                        "descripcion": "Base",
                        "activo": True,
                    },
                    {
                        "id": 2,
                        "codigo": "VOL",
                        "nombre": "Volumen",
                        "descripcion": "Nueva",
                        "activo": True,
                    },
                ],
                "proveedores": [
                    {
                        "id": 1,
                        "nombre": "Proveedor Uno",
                        "ruc": "20999999999",
                        "direccion": "Av. Nueva 123",
                    },
                    {
                        "id": 2,
                        "nombre": "Proveedor Dos",
                        "ruc": "20888888888",
                        "direccion": "Jr. Comercio",
                    },
                ],
                "ubicaciones_cliente": [
                    {
                        "id": 1,
                        "cliente": 1,
                        "nombre": "Taller Norte",
                        "direccion": "Zona 1",
                    }
                ],
                "unidades_medida": [
                    {
                        "id": 1,
                        "nombre": "METRO",
                        "simbolo": "m",
                        "dimension": 1,
                        "es_base": True,
                        "activo": True,
                    },
                    {
                        "id": 2,
                        "nombre": "LITRO",
                        "simbolo": "L",
                        "dimension": 2,
                        "es_base": True,
                        "activo": True,
                    },
                ],
                "items": [
                    {
                        "id": 1,
                        "codigo": "ITM-001",
                        "nombre": "Manguera Premium",
                        "tipo_insumo": "CONSUMIBLE",
                        "dimension": 1,
                        "unidad_medida": 1,
                        "favorito": True,
                        "volvo": False,
                        "ultimo_correlativo": 7,
                    },
                    {
                        "id": 2,
                        "codigo": "ITM-002",
                        "nombre": "Aceite Hidraulico",
                        "tipo_insumo": "CONSUMIBLE",
                        "dimension": 2,
                        "unidad_medida": 2,
                        "favorito": False,
                        "volvo": True,
                        "ultimo_correlativo": 0,
                    },
                ],
            }
        }
        upload = SimpleUploadedFile(
            "catalogos.json",
            json.dumps(payload).encode("utf-8"),
            content_type="application/json",
        )

        response = self.client.post(
            "/api/catalogo-sync/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["summary"]["processed"], 13)
        self.assertEqual(response.data["summary"]["created"], 7)
        self.assertEqual(response.data["summary"]["updated"], 3)
        self.assertEqual(response.data["summary"]["unchanged"], 3)

        self.maquinaria.refresh_from_db()
        self.proveedor.refresh_from_db()
        self.item.refresh_from_db()

        self.assertEqual(self.maquinaria.nombre, "Excavadora 320")
        self.assertEqual(str(self.maquinaria.gasto), "22.50")
        self.assertEqual(self.proveedor.direccion, "Av. Nueva 123")
        self.assertEqual(self.item.nombre, "Manguera Premium")
        self.assertTrue(self.item.favorito)
        self.assertEqual(self.item.ultimo_correlativo, 7)

        self.assertTrue(Maquinaria.objects.filter(pk=2).exists())
        self.assertTrue(Cliente.objects.filter(pk=2).exists())
        self.assertTrue(Dimension.objects.filter(pk=2).exists())
        self.assertTrue(Proveedor.objects.filter(pk=2).exists())
        self.assertTrue(UbicacionCliente.objects.filter(pk=1).exists())
        self.assertTrue(UnidadMedida.objects.filter(pk=2).exists())
        self.assertTrue(Item.objects.filter(pk=2).exists())

    def test_import_selected_table_from_csv(self):
        csv_content = (
            "id,codigo_maquina,nombre,descripcion,observacion,gasto\n"
            "1,MQ-01,Excavadora 320,Actualizada,Operativa,22.50\n"
            "2,MQ-02,Motoniveladora,,,0.00\n"
        )
        upload = SimpleUploadedFile(
            "maquinarias.csv",
            csv_content.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/catalogo-sync/",
            {"table": "maquinarias", "file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["summary"]["processed"], 2)
        self.assertEqual(response.data["summary"]["created"], 1)
        self.assertEqual(response.data["summary"]["updated"], 1)
        self.assertEqual(response.data["summary"]["tables"][0]["key"], "maquinarias")

        self.maquinaria.refresh_from_db()
        self.assertEqual(self.maquinaria.nombre, "Excavadora 320")
        self.assertEqual(str(self.maquinaria.gasto), "22.50")
        self.assertTrue(Maquinaria.objects.filter(pk=2).exists())

    def test_import_selected_table_from_xlsx(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["id", "nombre", "ruc"])
        worksheet.append([1, "Cliente Uno Premium", "20111111111"])
        worksheet.append([2, "Cliente Dos", "20222222222"])

        output = BytesIO()
        workbook.save(output)
        upload = SimpleUploadedFile(
            "clientes.xlsx",
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        response = self.client.post(
            "/api/catalogo-sync/",
            {"table": "clientes", "file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["summary"]["processed"], 2)
        self.assertEqual(response.data["summary"]["created"], 1)
        self.assertEqual(response.data["summary"]["updated"], 1)

        self.cliente.refresh_from_db()
        self.assertEqual(self.cliente.nombre, "Cliente Uno Premium")
        self.assertTrue(Cliente.objects.filter(pk=2, nombre="Cliente Dos").exists())


class ItemYCompraLegacyTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="admin-legacy",
            password="secret123",
        )
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        self.client.force_authenticate(user=self.user)

        self.dimension_unidad = Dimension.objects.create(
            codigo="UNIDAD",
            nombre="Unidad",
            descripcion="Conteo unitario",
            activo=True,
        )
        self.unidad_cantidad = UnidadMedida.objects.create(
            nombre="Cantidad",
            simbolo="und",
            dimension=self.dimension_unidad,
            es_base=True,
            activo=True,
        )

    def test_crear_herramienta_asigna_unidad_por_defecto(self):
        response = self.client.post(
            "/api/items/",
            {
                "codigo": "TOOL-001",
                "nombre": "Llave Stilson",
                "tipo_insumo": Item.TipoInsumo.HERRAMIENTA,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)

        item = Item.objects.get(pk=response.data["id"])
        self.assertEqual(item.dimension_id, self.dimension_unidad.id)
        self.assertEqual(item.unidad_medida_id, self.unidad_cantidad.id)

    def test_compra_herramienta_legacy_sanea_configuracion_y_registra_unidades(self):
        item = Item.objects.create(
            codigo="TOOL-LEG-001",
            nombre="Martillo",
            tipo_insumo=Item.TipoInsumo.HERRAMIENTA,
            dimension=None,
            unidad_medida=None,
        )

        response = self.client.post(
            "/api/compras/batch/",
            {
                "fecha": "2026-05-11",
                "moneda": "PEN",
                "items": [
                    {
                        "item": item.id,
                        "cantidad": 2,
                        "unidad_medida": None,
                        "tipo_registro": "VALOR_UNITARIO",
                        "monto": "25.00",
                        "moneda": "PEN",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)

        item.refresh_from_db()
        self.assertEqual(item.dimension_id, self.dimension_unidad.id)
        self.assertEqual(item.unidad_medida_id, self.unidad_cantidad.id)
        self.assertEqual(item.unidades.count(), 2)


class MaquinariaHorometroActualTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="admin-maq-horometro",
            password="secret123",
        )
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        self.client.force_authenticate(user=self.user)

        self.maquinaria = Maquinaria.objects.create(
            codigo_maquina="MQ-HM-01",
            nombre="Excavadora HM",
            descripcion="Prueba",
            observacion="",
            gasto="0.00",
        )

    def _crear_orden(self, *, fecha, horometro, hora_inicio=None, hora_fin=None):
        return OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            fecha=fecha,
            horometro=horometro,
            hora_inicio=hora_inicio,
            hora_fin=hora_fin,
            prioridad="REGULAR",
            lugar=OrdenTrabajo.Lugar.TALLER,
            observaciones="",
        )

    def test_lista_maquinaria_usa_horometro_manual_si_es_mas_reciente_que_la_ot(self):
        self._crear_orden(
            fecha=date(2026, 5, 10),
            horometro=Decimal("120.00"),
        )
        self.maquinaria.horometro_manual = Decimal("135.50")
        self.maquinaria.horometro_manual_actualizado_en = datetime(2026, 5, 11, 8, 0, 0)
        self.maquinaria.save(update_fields=["horometro_manual", "horometro_manual_actualizado_en"])

        response = self.client.get("/api/maquinarias/")

        self.assertEqual(response.status_code, 200, response.data)
        maquinaria = response.data[0]
        self.assertEqual(Decimal(str(maquinaria["horometro_actual"])), Decimal("135.50"))
        self.assertEqual(maquinaria["horometro_fuente"], "MANUAL")
        self.assertEqual(str(maquinaria["fecha_ultimo_horometro"]), "2026-05-11")

    def test_lista_maquinaria_usa_horometro_de_ot_si_la_ot_tiene_fecha_mas_reciente(self):
        self.maquinaria.horometro_manual = Decimal("135.50")
        self.maquinaria.horometro_manual_actualizado_en = datetime(2026, 5, 11, 8, 0, 0)
        self.maquinaria.save(update_fields=["horometro_manual", "horometro_manual_actualizado_en"])
        self._crear_orden(
            fecha=date(2026, 5, 12),
            horometro=Decimal("148.00"),
        )

        response = self.client.get("/api/maquinarias/")

        self.assertEqual(response.status_code, 200, response.data)
        maquinaria = response.data[0]
        self.assertEqual(Decimal(str(maquinaria["horometro_actual"])), Decimal("148.00"))
        self.assertEqual(maquinaria["horometro_fuente"], "ORDEN_TRABAJO")
        self.assertEqual(str(maquinaria["fecha_ultimo_horometro"]), "2026-05-12")

    def test_put_maquinaria_permite_editar_horometro_manual_y_actualiza_fecha_manual(self):
        response = self.client.put(
            f"/api/maquinarias/{self.maquinaria.id}/",
            {
                "codigo_maquina": self.maquinaria.codigo_maquina,
                "nombre": self.maquinaria.nombre,
                "descripcion": self.maquinaria.descripcion,
                "observacion": self.maquinaria.observacion,
                "gasto": "0.00",
                "horometro_manual": "222.25",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.maquinaria.refresh_from_db()
        self.assertEqual(self.maquinaria.horometro_manual, Decimal("222.25"))
        self.assertIsNotNone(self.maquinaria.horometro_manual_actualizado_en)
        self.assertEqual(Decimal(str(response.data["horometro_actual"])), Decimal("222.25"))
        self.assertEqual(response.data["horometro_fuente"], "MANUAL")


class OrdenTrabajoHistorialConsumibleHorometroTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="admin-horometro-cons",
            password="secret123",
        )
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        self.client.force_authenticate(user=self.user)

        self.dimension = Dimension.objects.create(
            codigo="VOL",
            nombre="Volumen",
            descripcion="Consumibles",
            activo=True,
        )
        self.unidad = UnidadMedida.objects.create(
            nombre="LITRO",
            simbolo="L",
            dimension=self.dimension,
            es_base=True,
            activo=True,
        )
        self.item = Item.objects.create(
            codigo="CONS-001",
            nombre="Aceite",
            tipo_insumo=Item.TipoInsumo.CONSUMIBLE,
            dimension=self.dimension,
            unidad_medida=self.unidad,
        )
        self.maquinaria = Maquinaria.objects.create(
            codigo_maquina="MQ-HT-01",
            nombre="Excavadora",
            descripcion="Prueba",
            observacion="",
            gasto="0.00",
        )
        self.trabajador = Trabajador.objects.create(
            nombres="Ana",
            apellidos="Perez",
            dni="12345678",
            puesto="Tecnico",
        )
        self.almacen = Almacen.objects.create(nombre="Principal")
        self.proveedor = Proveedor.objects.create(
            nombre="Proveedor HT",
            ruc="20123456789",
            direccion="Av. Test",
        )
        self.compra = Compra.objects.create(proveedor=self.proveedor)
        self.compra_detalle = CompraDetalle.objects.create(
            compra=self.compra,
            item=self.item,
            cantidad=100,
            unidad_medida=self.unidad,
            moneda=Compra.Moneda.PEN,
            valor_unitario="10.00",
        )

    def _crear_lote(self):
        return LoteConsumible.objects.create(
            compra_detalle=self.compra_detalle,
            item=self.item,
            cantidad_inicial=Decimal("100.000000"),
            cantidad_disponible=Decimal("100.000000"),
            unidad_medida=self.unidad,
            almacen=self.almacen,
        )

    def _crear_orden(self, fecha, horometro=None):
        orden = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            fecha=fecha,
            horometro=horometro,
            prioridad="REGULAR",
            lugar=OrdenTrabajo.Lugar.TALLER,
            observaciones="",
        )
        orden.tecnicos.add(self.trabajador)
        return orden

    def _crear_actividad_registrada(self, orden):
        return ActividadTrabajo.objects.create(
            orden=orden,
            tipo_actividad=ActividadTrabajo.TipoActividad.MANTENIMIENTO,
            tipo_mantenimiento=ActividadTrabajo.TipoMantenimiento.PREVENTIVO,
            subtipo=ActividadTrabajo.SubTipo.PM1,
            descripcion="Registro",
            es_planificada=False,
        )

    def _actualizar_horometro(self, orden, horometro):
        serializer = OrdenTrabajoSerializer(
            instance=orden,
            data={"horometro": str(horometro)},
            partial=True,
        )
        self.assertTrue(serializer.is_valid(), serializer.errors)
        serializer.save()

    def test_movimiento_consumible_registrado_crea_historial_en_maquinaria(self):
        lote = self._crear_lote()
        lote.cantidad_disponible = Decimal("10.000000")
        lote.save(update_fields=["cantidad_disponible"])

        historial_tecnico = HistorialConsumible.objects.create(
            lote=lote,
            item=self.item,
            cantidad=Decimal("10.000000"),
            unidad_medida=self.unidad,
            trabajador=self.trabajador,
        )

        orden = self._crear_orden(date(2026, 5, 2))
        actividad = self._crear_actividad_registrada(orden)

        response = self.client.post(
            "/api/movimientos-consumible/",
            {
                "actividad": actividad.id,
                "item": self.item.id,
                "cantidad": "5",
                "tecnico": self.trabajador.id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)

        lote.refresh_from_db()
        historial_tecnico.refresh_from_db()
        historial_maquinaria = HistorialConsumible.objects.get(
            orden_trabajo=orden,
            maquinaria=self.maquinaria,
            item=self.item,
        )
        historial_tecnico_restante = HistorialConsumible.objects.get(
            trabajador=self.trabajador,
            item=self.item,
            fecha_fin__isnull=True,
        )

        self.assertIsNone(historial_maquinaria.trabajador)
        self.assertEqual(historial_maquinaria.maquinaria_id, self.maquinaria.id)
        self.assertIsNone(historial_maquinaria.horometro_inicio)
        self.assertEqual(historial_maquinaria.cantidad, Decimal("5.000000"))
        self.assertEqual(lote.cantidad_disponible, Decimal("10.000000"))
        self.assertFalse(historial_tecnico.fecha_fin is None)
        self.assertEqual(historial_tecnico.cantidad, Decimal("5.000000"))
        self.assertEqual(historial_tecnico_restante.cantidad, Decimal("5.000000"))

    def test_horometros_solo_afectan_historiales_consumibles_de_maquinaria_registrada(self):
        orden_muy_anterior = self._crear_orden(date(2026, 5, 1), horometro=Decimal("80.00"))
        orden_anterior = self._crear_orden(date(2026, 5, 2), horometro=Decimal("100.00"))
        orden_actual = self._crear_orden(date(2026, 5, 3))

        actividad_muy_anterior = self._crear_actividad_registrada(orden_muy_anterior)
        actividad_anterior = self._crear_actividad_registrada(orden_anterior)
        actividad_actual = self._crear_actividad_registrada(orden_actual)

        MovimientoConsumible.objects.create(
            actividad=actividad_muy_anterior,
            item=self.item,
            cantidad=Decimal("2.000000"),
            unidad_medida=self.unidad,
            tecnico=self.trabajador,
        )
        MovimientoConsumible.objects.create(
            actividad=actividad_anterior,
            item=self.item,
            cantidad=Decimal("5.000000"),
            unidad_medida=self.unidad,
            tecnico=self.trabajador,
        )
        MovimientoConsumible.objects.create(
            actividad=actividad_actual,
            item=self.item,
            cantidad=Decimal("3.000000"),
            unidad_medida=self.unidad,
            tecnico=self.trabajador,
        )

        historial_muy_anterior = HistorialConsumible.objects.create(
            lote=self._crear_lote(),
            item=self.item,
            cantidad=Decimal("2.000000"),
            unidad_medida=self.unidad,
            maquinaria=self.maquinaria,
            orden_trabajo=orden_muy_anterior,
            fecha_fin=timezone.now(),
            horometro_inicio=Decimal("80.00"),
        )
        historial_anterior_1 = HistorialConsumible.objects.create(
            lote=self._crear_lote(),
            item=self.item,
            cantidad=Decimal("4.000000"),
            unidad_medida=self.unidad,
            maquinaria=self.maquinaria,
            orden_trabajo=orden_anterior,
            fecha_fin=timezone.now(),
            horometro_inicio=Decimal("100.00"),
        )
        historial_anterior_2 = HistorialConsumible.objects.create(
            lote=self._crear_lote(),
            item=self.item,
            cantidad=Decimal("1.000000"),
            unidad_medida=self.unidad,
            maquinaria=self.maquinaria,
            orden_trabajo=orden_anterior,
            fecha_fin=timezone.now(),
            horometro_inicio=Decimal("100.00"),
        )
        historial_anterior_tecnico = HistorialConsumible.objects.create(
            lote=self._crear_lote(),
            item=self.item,
            cantidad=Decimal("1.500000"),
            unidad_medida=self.unidad,
            trabajador=self.trabajador,
            orden_trabajo=orden_anterior,
            fecha_fin=timezone.now(),
        )
        historial_actual_maquinaria = HistorialConsumible.objects.create(
            lote=self._crear_lote(),
            item=self.item,
            cantidad=Decimal("3.000000"),
            unidad_medida=self.unidad,
            maquinaria=self.maquinaria,
            orden_trabajo=orden_actual,
        )
        historial_actual_tecnico = HistorialConsumible.objects.create(
            lote=self._crear_lote(),
            item=self.item,
            cantidad=Decimal("2.000000"),
            unidad_medida=self.unidad,
            trabajador=self.trabajador,
            orden_trabajo=orden_actual,
        )

        self._actualizar_horometro(orden_actual, Decimal("150.00"))

        historial_muy_anterior.refresh_from_db()
        historial_anterior_1.refresh_from_db()
        historial_anterior_2.refresh_from_db()
        historial_anterior_tecnico.refresh_from_db()
        historial_actual_maquinaria.refresh_from_db()
        historial_actual_tecnico.refresh_from_db()

        self.assertIsNone(historial_muy_anterior.horometro_fin)
        self.assertEqual(historial_anterior_1.horometro_fin, Decimal("150.00"))
        self.assertEqual(historial_anterior_2.horometro_fin, Decimal("150.00"))
        self.assertIsNone(historial_anterior_tecnico.horometro_fin)
        self.assertEqual(historial_actual_maquinaria.horometro_inicio, Decimal("150.00"))
        self.assertIsNone(historial_actual_tecnico.horometro_inicio)

    def test_consumible_registrado_autocompleta_horometro_fin_en_historial_previo_abierto(self):
        lote = self._crear_lote()
        historial_tecnico = HistorialConsumible.objects.create(
            lote=lote,
            item=self.item,
            cantidad=Decimal("10.000000"),
            unidad_medida=self.unidad,
            trabajador=self.trabajador,
        )
        orden_anterior = self._crear_orden(date(2026, 5, 5), horometro=Decimal("100.00"))
        HistorialConsumible.objects.create(
            lote=self._crear_lote(),
            item=self.item,
            cantidad=Decimal("4.000000"),
            unidad_medida=self.unidad,
            maquinaria=self.maquinaria,
            orden_trabajo=orden_anterior,
            horometro_inicio=Decimal("100.00"),
        )
        orden_actual = self._crear_orden(date(2026, 5, 5), horometro=Decimal("150.00"))
        actividad_actual = self._crear_actividad_registrada(orden_actual)

        response = self.client.post(
            "/api/movimientos-consumible/",
            {
                "actividad": actividad_actual.id,
                "item": self.item.id,
                "cantidad": "5",
                "tecnico": self.trabajador.id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)

        historial_tecnico.refresh_from_db()
        historial_anterior = HistorialConsumible.objects.get(
            orden_trabajo=orden_anterior,
            maquinaria=self.maquinaria,
            item=self.item,
        )
        historial_actual = HistorialConsumible.objects.get(
            orden_trabajo=orden_actual,
            maquinaria=self.maquinaria,
            item=self.item,
        )

        self.assertEqual(historial_anterior.horometro_fin, Decimal("150.00"))
        self.assertIsNone(historial_anterior.fecha_fin)
        self.assertEqual(historial_actual.horometro_inicio, Decimal("150.00"))
        self.assertEqual(historial_actual.cantidad, Decimal("5.000000"))
        self.assertEqual(historial_tecnico.cantidad, Decimal("5.000000"))


class OrdenTrabajoHistorialRepuestoHorometroTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="admin-ht-rep",
            password="secret123",
        )
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        self.client.force_authenticate(user=self.user)

        self.maquinaria = Maquinaria.objects.create(
            codigo_maquina="MQ-HT-REP-01",
            nombre="Retroexcavadora",
            descripcion="Prueba",
            observacion="",
            gasto="0.00",
        )
        self.trabajador = Trabajador.objects.create(
            nombres="Jorge",
            apellidos="Rivas",
            dni="11223344",
            puesto="Tecnico",
        )
        self.item = Item.objects.create(
            codigo="REP-HT-001",
            nombre="Filtro de aceite",
            tipo_insumo=Item.TipoInsumo.REPUESTO,
        )

    def _crear_orden(self, fecha, horometro=None):
        orden = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            fecha=fecha,
            horometro=horometro,
            prioridad="REGULAR",
            lugar=OrdenTrabajo.Lugar.TALLER,
            observaciones="",
        )
        orden.tecnicos.add(self.trabajador)
        return orden

    def _crear_actividad_registrada(self, orden):
        return ActividadTrabajo.objects.create(
            orden=orden,
            tipo_actividad=ActividadTrabajo.TipoActividad.MANTENIMIENTO,
            tipo_mantenimiento=ActividadTrabajo.TipoMantenimiento.PREVENTIVO,
            subtipo=ActividadTrabajo.SubTipo.PM1,
            descripcion="Registro repuesto",
            es_planificada=False,
        )

    def _crear_unidad_asignada_a_tecnico(self):
        unidad = ItemUnidad.objects.create(
            item=self.item,
            estado=ItemUnidad.Estado.NUEVO,
        )
        HistorialUbicacionItem.objects.create(
            item_unidad=unidad,
            trabajador=self.trabajador,
            estado=unidad.estado,
        )
        return unidad

    def _actualizar_horometro(self, orden, horometro):
        serializer = OrdenTrabajoSerializer(
            instance=orden,
            data={"horometro": str(horometro)},
            partial=True,
        )
        self.assertTrue(serializer.is_valid(), serializer.errors)
        serializer.save()

    def test_movimiento_repuesto_registrado_autocompleta_horometro_fin_previo_si_la_ot_ya_tiene_horometro(self):
        unidad_anterior = self._crear_unidad_asignada_a_tecnico()
        unidad_actual = self._crear_unidad_asignada_a_tecnico()

        orden_anterior = self._crear_orden(date(2026, 5, 5), horometro=Decimal("100.00"))
        actividad_anterior = self._crear_actividad_registrada(orden_anterior)
        response_anterior = self.client.post(
            "/api/movimientos-repuesto/",
            {
                "actividad": actividad_anterior.id,
                "item_unidad": unidad_anterior.id,
                "tecnico": self.trabajador.id,
            },
            format="json",
        )
        self.assertEqual(response_anterior.status_code, 201, response_anterior.data)

        historial_anterior = HistorialUbicacionItem.objects.get(
            orden_trabajo=orden_anterior,
            maquinaria=self.maquinaria,
            item_unidad=unidad_anterior,
        )
        self.assertEqual(historial_anterior.horometro_inicio, Decimal("100.00"))
        self.assertIsNone(historial_anterior.horometro_fin)

        orden_actual = self._crear_orden(date(2026, 5, 6), horometro=Decimal("150.00"))
        actividad_actual = self._crear_actividad_registrada(orden_actual)
        response_actual = self.client.post(
            "/api/movimientos-repuesto/",
            {
                "actividad": actividad_actual.id,
                "item_unidad": unidad_actual.id,
                "tecnico": self.trabajador.id,
            },
            format="json",
        )
        self.assertEqual(response_actual.status_code, 201, response_actual.data)

        historial_anterior.refresh_from_db()
        historial_actual = HistorialUbicacionItem.objects.get(
            orden_trabajo=orden_actual,
            maquinaria=self.maquinaria,
            item_unidad=unidad_actual,
        )

        self.assertEqual(historial_anterior.horometro_fin, Decimal("150.00"))
        self.assertEqual(historial_actual.horometro_inicio, Decimal("150.00"))

    def test_actualizar_horometro_de_ot_posterior_completa_horometro_fin_previo_de_repuesto(self):
        unidad_anterior = self._crear_unidad_asignada_a_tecnico()
        unidad_actual = self._crear_unidad_asignada_a_tecnico()

        orden_anterior = self._crear_orden(date(2026, 5, 10), horometro=Decimal("200.00"))
        actividad_anterior = self._crear_actividad_registrada(orden_anterior)
        response_anterior = self.client.post(
            "/api/movimientos-repuesto/",
            {
                "actividad": actividad_anterior.id,
                "item_unidad": unidad_anterior.id,
                "tecnico": self.trabajador.id,
            },
            format="json",
        )
        self.assertEqual(response_anterior.status_code, 201, response_anterior.data)

        orden_actual = self._crear_orden(date(2026, 5, 11))
        actividad_actual = self._crear_actividad_registrada(orden_actual)
        response_actual = self.client.post(
            "/api/movimientos-repuesto/",
            {
                "actividad": actividad_actual.id,
                "item_unidad": unidad_actual.id,
                "tecnico": self.trabajador.id,
            },
            format="json",
        )
        self.assertEqual(response_actual.status_code, 201, response_actual.data)

        historial_anterior = HistorialUbicacionItem.objects.get(
            orden_trabajo=orden_anterior,
            maquinaria=self.maquinaria,
            item_unidad=unidad_anterior,
        )
        historial_actual = HistorialUbicacionItem.objects.get(
            orden_trabajo=orden_actual,
            maquinaria=self.maquinaria,
            item_unidad=unidad_actual,
        )

        self.assertIsNone(historial_anterior.horometro_fin)
        self.assertIsNone(historial_actual.horometro_inicio)

        self._actualizar_horometro(orden_actual, Decimal("260.00"))

        historial_anterior.refresh_from_db()
        historial_actual.refresh_from_db()

        self.assertEqual(historial_anterior.horometro_fin, Decimal("260.00"))
        self.assertEqual(historial_actual.horometro_inicio, Decimal("260.00"))


class MovimientoConsumiblePlanificadoTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="admin-plan-cons",
            password="secret123",
        )
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        self.client.force_authenticate(user=self.user)

        self.dimension = Dimension.objects.create(
            codigo="VOL-PLAN",
            nombre="Volumen plan",
            descripcion="Consumibles planificados",
            activo=True,
        )
        self.unidad = UnidadMedida.objects.create(
            nombre="LITRO-PLAN",
            simbolo="L",
            dimension=self.dimension,
            es_base=True,
            activo=True,
        )
        self.item = Item.objects.create(
            codigo="CONS-PLAN-001",
            nombre="Aceite planificado",
            tipo_insumo=Item.TipoInsumo.CONSUMIBLE,
            dimension=self.dimension,
            unidad_medida=self.unidad,
        )
        self.maquinaria = Maquinaria.objects.create(
            codigo_maquina="MQ-PLAN-01",
            nombre="Cargador frontal",
            descripcion="Prueba",
            observacion="",
            gasto="0.00",
        )
        self.tecnico = Trabajador.objects.create(
            nombres="Luis",
            apellidos="Ramos",
            dni="87654321",
            puesto="Tecnico",
        )
        self.almacen = Almacen.objects.create(nombre="Almacen Planificado")
        self.proveedor = Proveedor.objects.create(
            nombre="Proveedor Plan",
            ruc="20987654321",
            direccion="Av. Plan",
        )
        self.compra = Compra.objects.create(proveedor=self.proveedor)
        self.compra_detalle = CompraDetalle.objects.create(
            compra=self.compra,
            item=self.item,
            cantidad=50,
            unidad_medida=self.unidad,
            moneda=Compra.Moneda.PEN,
            valor_unitario="12.00",
        )
        self.lote = LoteConsumible.objects.create(
            compra_detalle=self.compra_detalle,
            item=self.item,
            cantidad_inicial=Decimal("50.000000"),
            cantidad_disponible=Decimal("50.000000"),
            unidad_medida=self.unidad,
            almacen=self.almacen,
        )
        self.orden = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            fecha=date(2026, 5, 4),
            prioridad="REGULAR",
            lugar=OrdenTrabajo.Lugar.TALLER,
            observaciones="",
        )
        self.orden.tecnicos.add(self.tecnico)
        self.actividad = ActividadTrabajo.objects.create(
            orden=self.orden,
            tipo_actividad=ActividadTrabajo.TipoActividad.MANTENIMIENTO,
            tipo_mantenimiento=ActividadTrabajo.TipoMantenimiento.PREVENTIVO,
            subtipo=ActividadTrabajo.SubTipo.PM1,
            descripcion="Planificacion",
            es_planificada=True,
        )

    def test_movimiento_consumible_planificado_no_crea_historial(self):
        HistorialConsumible.objects.create(
            lote=self.lote,
            item=self.item,
            cantidad=Decimal("10.000000"),
            unidad_medida=self.unidad,
            trabajador=self.tecnico,
        )
        total_historiales_antes = HistorialConsumible.objects.count()

        response = self.client.post(
            "/api/movimientos-consumible/",
            {
                "actividad": self.actividad.id,
                "item": self.item.id,
                "cantidad": "5",
                "tecnico": self.tecnico.id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(HistorialConsumible.objects.count(), total_historiales_antes)
        self.assertEqual(MovimientoConsumible.objects.count(), 1)

        movimiento = MovimientoConsumible.objects.get()
        self.assertEqual(movimiento.tecnico_id, self.tecnico.id)
        self.assertEqual(response.data["tecnico"], self.tecnico.id)
        self.assertEqual(response.data["tecnico_nombre"], "Luis Ramos")


class MovimientoRepuestoPlanificadoTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="admin-plan-rep",
            password="secret123",
        )
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        self.client.force_authenticate(user=self.user)

        self.maquinaria = Maquinaria.objects.create(
            codigo_maquina="MQ-PLAN-REP-01",
            nombre="Excavadora plan",
            descripcion="Prueba",
            observacion="",
            gasto="0.00",
        )
        self.tecnico = Trabajador.objects.create(
            nombres="Mario",
            apellidos="Campos",
            dni="88776655",
            puesto="Tecnico",
        )
        self.almacen = Almacen.objects.create(nombre="Almacen Plan Repuesto")
        self.orden = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            fecha=date(2026, 5, 6),
            prioridad="REGULAR",
            lugar=OrdenTrabajo.Lugar.TALLER,
            observaciones="",
        )
        self.orden.tecnicos.add(self.tecnico)
        self.actividad_planificada = ActividadTrabajo.objects.create(
            orden=self.orden,
            tipo_actividad=ActividadTrabajo.TipoActividad.MANTENIMIENTO,
            tipo_mantenimiento=ActividadTrabajo.TipoMantenimiento.PREVENTIVO,
            subtipo=ActividadTrabajo.SubTipo.PM1,
            descripcion="Planificacion repuestos",
            es_planificada=True,
        )
        self.actividad_real = ActividadTrabajo.objects.create(
            orden=self.orden,
            tipo_actividad=ActividadTrabajo.TipoActividad.MANTENIMIENTO,
            tipo_mantenimiento=ActividadTrabajo.TipoMantenimiento.PREVENTIVO,
            subtipo=ActividadTrabajo.SubTipo.PM1,
            descripcion="Ejecucion repuestos",
            es_planificada=False,
        )
        self.item = Item.objects.create(
            codigo="REP-PLAN-001",
            nombre="Filtro planificado",
            tipo_insumo=Item.TipoInsumo.REPUESTO,
        )
        self.unidad = ItemUnidad.objects.create(
            item=self.item,
            estado=ItemUnidad.Estado.NUEVO,
        )
        HistorialUbicacionItem.objects.create(
            item_unidad=self.unidad,
            trabajador=self.tecnico,
            estado=self.unidad.estado,
        )
        self.unidad_alterna = ItemUnidad.objects.create(
            item=self.item,
            estado=ItemUnidad.Estado.NUEVO,
        )
        HistorialUbicacionItem.objects.create(
            item_unidad=self.unidad_alterna,
            trabajador=self.tecnico,
            estado=self.unidad_alterna.estado,
        )

    def test_movimiento_repuesto_planificado_no_crea_historial(self):
        total_historiales_antes = HistorialUbicacionItem.objects.count()

        response = self.client.post(
            "/api/movimientos-repuesto/",
            {
                "actividad": self.actividad_planificada.id,
                "item_unidad": self.unidad.id,
                "tecnico": self.tecnico.id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(HistorialUbicacionItem.objects.count(), total_historiales_antes)
        self.assertEqual(MovimientoRepuesto.objects.count(), 1)

        movimiento = MovimientoRepuesto.objects.get()
        self.assertEqual(movimiento.tecnico_id, self.tecnico.id)
        self.assertEqual(response.data["tecnico"], self.tecnico.id)
        self.assertEqual(response.data["tecnico_id"], self.tecnico.id)
        self.assertEqual(response.data["tecnico_nombre"], "Mario Campos")

    def test_movimiento_real_reutiliza_unidad_planificada_sin_historial_previsto(self):
        plan_response = self.client.post(
            "/api/movimientos-repuesto/",
            {
                "actividad": self.actividad_planificada.id,
                "item_unidad": self.unidad.id,
                "tecnico": self.tecnico.id,
            },
            format="json",
        )
        self.assertEqual(plan_response.status_code, 201, plan_response.data)

        historial_count_antes = HistorialUbicacionItem.objects.count()

        response = self.client.post(
            "/api/movimientos-repuesto/",
            {
                "actividad": self.actividad_real.id,
                "item_unidad": self.unidad_alterna.id,
                "tecnico": self.tecnico.id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)

        movimiento_real = MovimientoRepuesto.objects.get(actividad=self.actividad_real)
        self.assertEqual(movimiento_real.item_unidad_id, self.unidad.id)
        self.assertEqual(HistorialUbicacionItem.objects.count(), historial_count_antes + 1)

        historial_real = HistorialUbicacionItem.objects.filter(
            item_unidad=self.unidad,
            orden_trabajo=self.orden,
        ).order_by("-id").first()
        self.assertIsNotNone(historial_real)
        self.assertEqual(historial_real.maquinaria_id, self.maquinaria.id)
        self.assertIsNone(historial_real.trabajador_id)
        self.assertEqual(historial_real.estado, ItemUnidad.Estado.USADO)

        self.unidad.refresh_from_db()
        self.assertEqual(self.unidad.maquinaria_actual_id, self.maquinaria.id)
        self.assertIsNone(self.unidad.trabajador_actual_id)


class OrdenRequerimientoPermisosTests(APITestCase):
    def setUp(self):
        self.jefe_tecnicos_group = Group.objects.create(name="Jefe de Tecnicos")
        self.jefe_almacen_group = Group.objects.create(name="Jefe de Almaceneros")
        self.almacenero_group = Group.objects.create(name="Almacenero")
        self.tecnico_group = Group.objects.create(name="Tecnico")

        self.jefe_tecnicos_user = User.objects.create_user(
            username="jefe-tecnicos",
            password="secret123",
        )
        self.jefe_tecnicos_user.groups.add(self.jefe_tecnicos_group)

        self.almacen_user = User.objects.create_user(
            username="jefe-almacen",
            password="secret123",
        )
        self.almacen_user.groups.add(self.jefe_almacen_group)

        self.tecnico = Trabajador.objects.create(
            nombres="Pedro",
            apellidos="Quispe",
            dni="44556677",
            puesto="Tecnico",
        )
        self.tecnico_user = User.objects.create_user(
            username="tecnico-asignado",
            password="secret123",
        )
        self.tecnico_user.groups.add(self.tecnico_group)
        PerfilUsuario.objects.create(user=self.tecnico_user, trabajador=self.tecnico)

        self.dimension = Dimension.objects.create(
            codigo="VOL-REQ",
            nombre="Volumen requerimiento",
            descripcion="Base",
            activo=True,
        )
        self.unidad = UnidadMedida.objects.create(
            nombre="LITRO-REQ",
            simbolo="L",
            dimension=self.dimension,
            es_base=True,
            activo=True,
        )
        self.item = Item.objects.create(
            codigo="CONS-REQ-001",
            nombre="Aceite motor",
            tipo_insumo=Item.TipoInsumo.CONSUMIBLE,
            dimension=self.dimension,
            unidad_medida=self.unidad,
        )
        self.item_extra = Item.objects.create(
            codigo="CONS-REQ-002",
            nombre="Refrigerante",
            tipo_insumo=Item.TipoInsumo.CONSUMIBLE,
            dimension=self.dimension,
            unidad_medida=self.unidad,
        )
        self.almacen = Almacen.objects.create(nombre="Almacen Requerimientos")
        self.proveedor = Proveedor.objects.create(
            nombre="Proveedor Req",
            ruc="20123456789",
            direccion="Av. Requerimientos",
        )
        self.compra = Compra.objects.create(proveedor=self.proveedor)
        self.compra_detalle = CompraDetalle.objects.create(
            compra=self.compra,
            item=self.item,
            cantidad=10,
            unidad_medida=self.unidad,
            moneda=Compra.Moneda.PEN,
            valor_unitario="20.00",
        )
        self.compra_detalle_extra = CompraDetalle.objects.create(
            compra=self.compra,
            item=self.item_extra,
            cantidad=10,
            unidad_medida=self.unidad,
            moneda=Compra.Moneda.PEN,
            valor_unitario="15.00",
        )
        self.lote = LoteConsumible.objects.create(
            compra_detalle=self.compra_detalle,
            item=self.item,
            cantidad_inicial=Decimal("10.000000"),
            cantidad_disponible=Decimal("10.000000"),
            unidad_medida=self.unidad,
            almacen=self.almacen,
        )
        self.lote_extra = LoteConsumible.objects.create(
            compra_detalle=self.compra_detalle_extra,
            item=self.item_extra,
            cantidad_inicial=Decimal("10.000000"),
            cantidad_disponible=Decimal("10.000000"),
            unidad_medida=self.unidad,
            almacen=self.almacen,
        )
        HistorialConsumible.objects.create(
            lote=self.lote,
            item=self.item,
            cantidad=Decimal("10.000000"),
            unidad_medida=self.unidad,
            almacen=self.almacen,
        )
        HistorialConsumible.objects.create(
            lote=self.lote_extra,
            item=self.item_extra,
            cantidad=Decimal("10.000000"),
            unidad_medida=self.unidad,
            almacen=self.almacen,
        )

        self.maquinaria = Maquinaria.objects.create(
            codigo_maquina="MQ-REQ-01",
            nombre="Retroexcavadora",
            descripcion="Prueba",
            observacion="",
            gasto="0.00",
        )
        self.orden_trabajo = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            fecha=date(2026, 5, 12),
            prioridad="REGULAR",
            lugar=OrdenTrabajo.Lugar.TALLER,
            observaciones="",
        )
        self.orden_trabajo.tecnicos.add(self.tecnico)

    def _crear_requerimiento(self, tecnico_asignado=None):
        orden = OrdenRequerimiento.objects.create(
            trabajo=self.orden_trabajo,
            tecnico_asignado=tecnico_asignado,
            observaciones="Materiales para la OT",
            emitido_por=self.jefe_tecnicos_user,
        )
        OrdenRequerimientoDetalle.objects.create(
            orden_requerimiento=orden,
            item=self.item,
            cantidad=Decimal("2.000000"),
            unidad_medida=self.unidad,
            proveedor=self.proveedor,
        )
        return orden

    def test_jefe_tecnicos_puede_crear_orden_requerimiento(self):
        self.client.force_authenticate(user=self.jefe_tecnicos_user)

        response = self.client.post(
            "/api/ordenes-requerimiento/",
            {
                "trabajo": self.orden_trabajo.id,
                "tecnico_asignado": self.tecnico.id,
                "observaciones": "Solicitar aceite",
                "items": [
                    {
                        "item": self.item.id,
                        "cantidad": "2",
                        "proveedor": self.proveedor.id,
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["tecnico_asignado"], self.tecnico.id)
        self.assertEqual(OrdenRequerimiento.objects.count(), 1)
        self.assertEqual(response.data["items"][0]["unidad_medida"], self.unidad.id)
        self.assertEqual(response.data["items"][0]["unidad_medida_simbolo"], self.unidad.simbolo)
        self.assertEqual(
            OrdenRequerimientoDetalle.objects.get().unidad_medida_id,
            self.unidad.id,
        )

    def test_jefe_tecnicos_puede_listar_items_para_emitir_requerimientos(self):
        self.client.force_authenticate(user=self.jefe_tecnicos_user)

        response = self.client.get("/api/items/")

        self.assertEqual(response.status_code, 200, response.data)
        item_ids = [row["id"] for row in response.data]
        self.assertIn(self.item.id, item_ids)

    def test_jefe_tecnicos_puede_registrar_movimiento_consumible(self):
        actividad = ActividadTrabajo.objects.create(
            orden=self.orden_trabajo,
            tipo_actividad=ActividadTrabajo.TipoActividad.MANTENIMIENTO,
            tipo_mantenimiento=ActividadTrabajo.TipoMantenimiento.PREVENTIVO,
            subtipo=ActividadTrabajo.SubTipo.PM1,
            descripcion="Registro desde jefe de tecnicos",
            es_planificada=False,
        )
        HistorialConsumible.objects.create(
            lote=self.lote,
            item=self.item,
            cantidad=Decimal("5.000000"),
            unidad_medida=self.unidad,
            trabajador=self.tecnico,
        )

        self.client.force_authenticate(user=self.jefe_tecnicos_user)

        response = self.client.post(
            "/api/movimientos-consumible/",
            {
                "actividad": actividad.id,
                "item": self.item.id,
                "cantidad": "1",
                "tecnico": self.tecnico.id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["tecnico"], self.tecnico.id)

    def test_tecnico_asignado_puede_marcar_entregado(self):
        orden = self._crear_requerimiento(tecnico_asignado=self.tecnico)
        self.client.force_authenticate(user=self.tecnico_user)

        response = self.client.post(
            f"/api/ordenes-requerimiento/{orden.id}/cambiar_estado/",
            {"estado": OrdenRequerimiento.Estado.ENTREGADO},
            format="json",
        )

        self.assertEqual(response.status_code, 200, response.data)
        orden.refresh_from_db()
        self.assertEqual(orden.estado, OrdenRequerimiento.Estado.ENTREGADO)

    def test_tecnico_asignado_no_puede_marcar_sin_stock(self):
        orden = self._crear_requerimiento(tecnico_asignado=self.tecnico)
        self.client.force_authenticate(user=self.tecnico_user)

        response = self.client.post(
            f"/api/ordenes-requerimiento/{orden.id}/cambiar_estado/",
            {
                "estado": OrdenRequerimiento.Estado.SIN_STOCK,
                "detalle_id": orden.detalles.first().id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 403, response.data)
        orden.refresh_from_db()
        self.assertEqual(orden.estado, OrdenRequerimiento.Estado.POR_REVISAR)

    def test_almacen_puede_marcar_un_detalle_sin_stock(self):
        orden = self._crear_requerimiento(tecnico_asignado=self.tecnico)
        detalle = orden.detalles.first()
        self.client.force_authenticate(user=self.almacen_user)

        response = self.client.post(
            f"/api/ordenes-requerimiento/{orden.id}/cambiar_estado/",
            {
                "estado": OrdenRequerimiento.Estado.SIN_STOCK,
                "detalle_id": detalle.id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200, response.data)
        orden.refresh_from_db()
        detalle.refresh_from_db()
        self.assertEqual(orden.estado, OrdenRequerimiento.Estado.SIN_STOCK)
        self.assertTrue(detalle.sin_stock)

    def test_entrega_solo_asigna_items_disponibles(self):
        orden = OrdenRequerimiento.objects.create(
            trabajo=self.orden_trabajo,
            tecnico_asignado=self.tecnico,
            observaciones="Entrega parcial",
            emitido_por=self.jefe_tecnicos_user,
        )
        detalle_sin_stock = OrdenRequerimientoDetalle.objects.create(
            orden_requerimiento=orden,
            item=self.item,
            cantidad=Decimal("2.000000"),
            unidad_medida=self.unidad,
            proveedor=self.proveedor,
            sin_stock=True,
        )
        OrdenRequerimientoDetalle.objects.create(
            orden_requerimiento=orden,
            item=self.item_extra,
            cantidad=Decimal("1.000000"),
            unidad_medida=self.unidad,
            proveedor=self.proveedor,
        )

        self.client.force_authenticate(user=self.almacen_user)

        response = self.client.post(
            f"/api/ordenes-requerimiento/{orden.id}/cambiar_estado/",
            {"estado": OrdenRequerimiento.Estado.ENTREGADO},
            format="json",
        )

        self.assertEqual(response.status_code, 200, response.data)
        orden.refresh_from_db()
        detalle_sin_stock.refresh_from_db()
        self.assertEqual(orden.estado, OrdenRequerimiento.Estado.ENTREGADO)
        self.assertTrue(detalle_sin_stock.sin_stock)

        total_asignado_tecnico = (
            HistorialConsumible.objects.filter(
                trabajador=self.tecnico,
                orden_trabajo=self.orden_trabajo,
            )
            .aggregate(total=Sum("cantidad"))
            .get("total")
        )
        self.assertEqual(total_asignado_tecnico, Decimal("1.000000"))


class ActividadPlanificadaPermisosTests(APITestCase):
    def setUp(self):
        self.jefe_tecnicos_group = Group.objects.create(name="Jefe de Tecnicos")
        self.jefe_almacen_group = Group.objects.create(name="Jefe de Almaceneros")
        self.almacenero_group = Group.objects.create(name="Almacenero")

        self.jefe_tecnicos_user = User.objects.create_user(
            username="jefe-tecnicos-plan",
            password="secret123",
        )
        self.jefe_tecnicos_user.groups.add(self.jefe_tecnicos_group)

        self.jefe_almacen_user = User.objects.create_user(
            username="jefe-almacen-plan",
            password="secret123",
        )
        self.jefe_almacen_user.groups.add(self.jefe_almacen_group)

        self.almacenero_user = User.objects.create_user(
            username="almacenero-plan",
            password="secret123",
        )
        self.almacenero_user.groups.add(self.almacenero_group)

        self.maquinaria = Maquinaria.objects.create(
            codigo_maquina="MQ-ACT-01",
            nombre="Motoniveladora",
            descripcion="Prueba",
            observacion="",
            gasto="0.00",
        )
        self.orden = OrdenTrabajo.objects.create(
            maquinaria=self.maquinaria,
            fecha=date(2026, 5, 12),
            prioridad="REGULAR",
            lugar=OrdenTrabajo.Lugar.TALLER,
            observaciones="",
        )

    def test_jefe_tecnicos_puede_crear_actividad_planificada(self):
        self.client.force_authenticate(user=self.jefe_tecnicos_user)

        response = self.client.post(
            "/api/actividades/",
            {
                "orden": self.orden.id,
                "tipo_actividad": ActividadTrabajo.TipoActividad.REVISION,
                "descripcion": "Revision programada",
                "es_planificada": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertTrue(response.data["es_planificada"])

    def test_almacenero_no_puede_crear_actividad_planificada(self):
        self.client.force_authenticate(user=self.almacenero_user)

        response = self.client.post(
            "/api/actividades/",
            {
                "orden": self.orden.id,
                "tipo_actividad": ActividadTrabajo.TipoActividad.REVISION,
                "descripcion": "Revision programada",
                "es_planificada": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 403, response.data)

    def test_jefe_almacen_puede_crear_actividad_planificada(self):
        self.client.force_authenticate(user=self.jefe_almacen_user)

        response = self.client.post(
            "/api/actividades/",
            {
                "orden": self.orden.id,
                "tipo_actividad": ActividadTrabajo.TipoActividad.REVISION,
                "descripcion": "Revision programada",
                "es_planificada": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
