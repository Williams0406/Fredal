import csv
import json
from io import BytesIO, StringIO

from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.management.color import no_style
from django.db import IntegrityError, connection, transaction
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from rest_framework import status
from rest_framework.exceptions import ParseError, ValidationError
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import (
    ActividadTrabajo,
    Almacen,
    Cliente,
    Compra,
    CompraDetalle,
    Dimension,
    Item,
    ItemGrupo,
    ItemGrupoDetalle,
    ItemProveedor,
    Maquinaria,
    OrdenTrabajo,
    Proveedor,
    TecnicoAsignado,
    TipoCambioDiario,
    Trabajador,
    UbicacionCliente,
    UnidadMedida,
    UnidadRelacion,
)
from .permissions import CatalogoPermission


SUPPORTED_FORMATS = ("json", "csv", "xlsx")
XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

TABLE_CONFIGS = (
    {
        "key": "almacenes",
        "label": "Almacenes",
        "model": Almacen,
        "fields": ["id", "nombre"],
    },
    {
        "key": "maquinarias",
        "label": "Maquinarias",
        "model": Maquinaria,
        "fields": [
            "id",
            "codigo_maquina",
            "nombre",
            "descripcion",
            "observacion",
            "gasto",
        ],
    },
    {
        "key": "clientes",
        "label": "Clientes",
        "model": Cliente,
        "fields": ["id", "nombre", "ruc"],
    },
    {
        "key": "dimensiones",
        "label": "Dimensiones",
        "model": Dimension,
        "fields": ["id", "codigo", "nombre", "descripcion", "activo"],
    },
    {
        "key": "proveedores",
        "label": "Proveedores",
        "model": Proveedor,
        "fields": ["id", "nombre", "ruc", "direccion"],
    },
    {
        "key": "trabajadores",
        "label": "Trabajadores",
        "model": Trabajador,
        "fields": ["id", "codigo", "nombres", "apellidos", "dni", "puesto"],
    },
    {
        "key": "tipos_cambio",
        "label": "Tipos de cambio",
        "model": TipoCambioDiario,
        "fields": [
            "id",
            "fecha",
            "compra_usd",
            "venta_usd",
            "compra_eur",
            "venta_eur",
        ],
    },
    {
        "key": "ubicaciones_cliente",
        "label": "Ubicaciones de cliente",
        "model": UbicacionCliente,
        "fields": ["id", "cliente", "nombre", "direccion"],
        "foreign_keys": {
            "cliente": Cliente,
        },
    },
    {
        "key": "unidades_medida",
        "label": "Unidades de medida",
        "model": UnidadMedida,
        "fields": ["id", "nombre", "simbolo", "dimension", "es_base", "activo"],
        "foreign_keys": {
            "dimension": Dimension,
        },
    },
    {
        "key": "relaciones_unidad",
        "label": "Relaciones de unidad",
        "model": UnidadRelacion,
        "fields": [
            "id",
            "dimension",
            "unidad_base",
            "unidad_relacionada",
            "factor",
            "activo",
        ],
        "foreign_keys": {
            "dimension": Dimension,
            "unidad_base": UnidadMedida,
            "unidad_relacionada": UnidadMedida,
        },
    },
    {
        "key": "items",
        "label": "Items",
        "model": Item,
        "fields": [
            "id",
            "codigo",
            "nombre",
            "tipo_insumo",
            "dimension",
            "unidad_medida",
            "favorito",
            "volvo",
            "ultimo_correlativo",
        ],
        "foreign_keys": {
            "dimension": Dimension,
            "unidad_medida": UnidadMedida,
        },
    },
    {
        "key": "item_proveedores",
        "label": "Item Proveedores",
        "model": ItemProveedor,
        "fields": ["id", "item", "proveedor", "precio", "moneda"],
        "foreign_keys": {
            "item": Item,
            "proveedor": Proveedor,
        },
    },
    {
        "key": "item_grupos",
        "label": "Grupos de items",
        "model": ItemGrupo,
        "fields": ["id", "nombre"],
    },
    {
        "key": "item_grupo_detalles",
        "label": "Detalles de grupo de items",
        "model": ItemGrupoDetalle,
        "fields": ["id", "grupo", "item", "cantidad", "unidad_medida"],
        "foreign_keys": {
            "grupo": ItemGrupo,
            "item": Item,
            "unidad_medida": UnidadMedida,
        },
    },
    {
        "key": "compras",
        "label": "Compras",
        "model": Compra,
        "fields": [
            "id",
            "tipo_comprobante",
            "codigo_comprobante",
            "proveedor",
            "moneda",
            "fecha",
        ],
        "foreign_keys": {
            "proveedor": Proveedor,
        },
    },
    {
        "key": "compra_detalles",
        "label": "Detalles de compra",
        "model": CompraDetalle,
        "fields": [
            "id",
            "compra",
            "item",
            "cantidad",
            "unidad_medida",
            "moneda",
            "valor_unitario",
        ],
        "foreign_keys": {
            "compra": Compra,
            "item": Item,
            "unidad_medida": UnidadMedida,
        },
    },
    {
        "key": "ordenes_trabajo",
        "label": "Ordenes de trabajo",
        "model": OrdenTrabajo,
        "fields": [
            "id",
            "codigo_orden",
            "maquinaria",
            "fecha",
            "hora_inicio",
            "hora_fin",
            "horometro",
            "prioridad",
            "lugar",
            "ubicacion_detalle",
            "estado_equipo",
            "estatus",
            "observaciones",
        ],
        "foreign_keys": {
            "maquinaria": Maquinaria,
        },
    },
    {
        "key": "tecnicos_asignados",
        "label": "Tecnicos asignados",
        "model": TecnicoAsignado,
        "fields": ["id", "orden", "tecnico"],
        "foreign_keys": {
            "orden": OrdenTrabajo,
            "tecnico": Trabajador,
        },
    },
    {
        "key": "actividades_trabajo",
        "label": "Actividades de trabajo",
        "model": ActividadTrabajo,
        "fields": [
            "id",
            "orden",
            "tipo_actividad",
            "tipo_mantenimiento",
            "subtipo",
            "descripcion",
            "es_planificada",
        ],
        "foreign_keys": {
            "orden": OrdenTrabajo,
        },
    },
)

TABLE_CONFIG_MAP = {config["key"]: config for config in TABLE_CONFIGS}


class CatalogoSyncView(APIView):
    permission_classes = [CatalogoPermission]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get(self, request):
        if self._is_truthy(request.query_params.get("metadata")):
            return Response(self._build_metadata())

        table_key = request.query_params.get("table")
        export_format = (request.query_params.get("file_format") or "json").lower()

        if table_key:
            config = self._get_table_config(table_key)
            if export_format not in SUPPORTED_FORMATS:
                raise ValidationError(
                    {
                        "detail": (
                            "Formato no compatible. Usa json, csv o xlsx."
                        )
                    }
                )
            return self._export_single_table(config, export_format)

        return self._export_bundle_json()

    def post(self, request):
        table_key = request.data.get("table") or request.query_params.get("table")

        if table_key:
            config = self._get_table_config(table_key)
            rows = self._parse_selected_table_rows(request, config)
            summary = self._import_tables({config["key"]: rows})
        else:
            payload = self._parse_payload(request)
            tables_payload = self._extract_tables(payload)
            summary = self._import_tables(tables_payload)

        return Response(
            {
                "message": "Importacion completada correctamente.",
                "summary": summary,
            },
            status=status.HTTP_200_OK,
        )

    def _build_metadata(self):
        return {
            "formats": list(SUPPORTED_FORMATS),
            "tables": [
                {
                    "key": config["key"],
                    "label": config["label"],
                    "fields": config["fields"],
                    "record_count": config["model"].objects.count(),
                }
                for config in TABLE_CONFIGS
            ],
        }

    def _export_bundle_json(self):
        generated_at = timezone.now()
        if timezone.is_aware(generated_at):
            generated_at = timezone.localtime(generated_at)
        tables = {}
        record_counts = {}

        for config in TABLE_CONFIGS:
            records = self._get_table_records(config)
            tables[config["key"]] = records
            record_counts[config["key"]] = len(records)

        payload = {
            "meta": {
                "format": "fredal.catalog-sync",
                "version": 2,
                "generated_at": generated_at.isoformat(),
                "table_order": [config["key"] for config in TABLE_CONFIGS],
                "record_counts": record_counts,
            },
            "tables": tables,
        }

        response = JsonResponse(
            payload,
            json_dumps_params={
                "ensure_ascii": False,
                "indent": 2,
            },
        )
        response["Content-Disposition"] = (
            "attachment; "
            f'filename="catalogos-sync-{generated_at.strftime("%Y%m%d-%H%M%S")}.json"'
        )
        return response

    def _export_single_table(self, config, export_format):
        generated_at = timezone.now()
        if timezone.is_aware(generated_at):
            generated_at = timezone.localtime(generated_at)
        records = self._get_table_records(config)
        filename = (
            f'{config["key"]}-{generated_at.strftime("%Y%m%d-%H%M%S")}.{export_format}'
        )

        if export_format == "json":
            payload = {
                "meta": {
                    "format": "fredal.table-sync",
                    "version": 2,
                    "generated_at": generated_at.isoformat(),
                },
                "table": config["key"],
                "label": config["label"],
                "fields": config["fields"],
                "rows": records,
            }
            response = JsonResponse(
                payload,
                json_dumps_params={
                    "ensure_ascii": False,
                    "indent": 2,
                },
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        if export_format == "csv":
            csv_buffer = StringIO()
            writer = csv.DictWriter(csv_buffer, fieldnames=config["fields"])
            writer.writeheader()

            for row in records:
                writer.writerow(
                    {
                        field_name: self._serialize_tabular_value(row.get(field_name))
                        for field_name in config["fields"]
                    }
                )

            response = HttpResponse(
                "\ufeff" + csv_buffer.getvalue(),
                content_type="text/csv; charset=utf-8",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self._sheet_title(config["label"])
        worksheet.append(config["fields"])

        for row in records:
            worksheet.append(
                [
                    self._serialize_tabular_value(row.get(field_name))
                    for field_name in config["fields"]
                ]
            )

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type=XLSX_CONTENT_TYPE,
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _get_table_records(self, config):
        return list(
            config["model"].objects.order_by("id").values(*config["fields"])
        )

    def _parse_selected_table_rows(self, request, config):
        uploaded_file = request.FILES.get("file") or request.FILES.get("archivo")

        if uploaded_file:
            file_format = self._detect_uploaded_format(uploaded_file.name)

            if file_format == "csv":
                return self._parse_csv_file(uploaded_file)

            if file_format == "xlsx":
                return self._parse_xlsx_file(uploaded_file)

            payload = self._parse_uploaded_json(uploaded_file)
            return self._extract_rows_for_selected_table(payload, config)

        payload = self._parse_payload(request)
        return self._extract_rows_for_selected_table(payload, config)

    def _parse_payload(self, request):
        uploaded_file = request.FILES.get("file") or request.FILES.get("archivo")
        if uploaded_file:
            return self._parse_uploaded_json(uploaded_file)

        payload = request.data
        if hasattr(payload, "dict"):
            payload = payload.dict()

        if isinstance(payload, dict) and "payload" in payload:
            raw_payload = payload["payload"]
            if isinstance(raw_payload, str):
                try:
                    return json.loads(raw_payload)
                except json.JSONDecodeError as exc:
                    raise ParseError(
                        "El campo payload no contiene un JSON valido."
                    ) from exc

        if isinstance(payload, str):
            try:
                return json.loads(payload)
            except json.JSONDecodeError as exc:
                raise ParseError(
                    "El cuerpo de la solicitud no contiene un JSON valido."
                ) from exc

        if not isinstance(payload, dict):
            raise ParseError("No se encontro informacion valida para importar.")

        return payload

    def _parse_uploaded_json(self, uploaded_file):
        try:
            raw_payload = uploaded_file.read().decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ParseError("El archivo debe estar codificado en UTF-8.") from exc

        try:
            return json.loads(raw_payload)
        except json.JSONDecodeError as exc:
            raise ParseError(
                "El archivo seleccionado no contiene un JSON valido."
            ) from exc

    def _parse_csv_file(self, uploaded_file):
        try:
            raw_content = uploaded_file.read().decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ParseError("El archivo CSV debe estar codificado en UTF-8.") from exc

        reader = csv.DictReader(StringIO(raw_content))
        if not reader.fieldnames:
            raise ParseError(
                "El archivo CSV debe incluir encabezados en la primera fila."
            )

        rows = []
        for row in reader:
            normalized_row = self._normalize_tabular_row(row)
            if normalized_row is not None:
                rows.append(normalized_row)

        return rows

    def _parse_xlsx_file(self, uploaded_file):
        try:
            workbook = load_workbook(
                filename=BytesIO(uploaded_file.read()),
                data_only=True,
            )
        except InvalidFileException as exc:
            raise ParseError("El archivo Excel no es valido.") from exc
        except Exception as exc:
            raise ParseError("No se pudo leer el archivo Excel.") from exc

        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)

        try:
            headers_row = next(iterator)
        except StopIteration as exc:
            raise ParseError("El archivo Excel no contiene datos.") from exc

        headers = [
            str(header).strip() if header is not None else ""
            for header in headers_row
        ]
        if not any(headers):
            raise ParseError(
                "La primera fila del archivo Excel debe contener encabezados."
            )

        rows = []
        for values in iterator:
            raw_row = {
                headers[index]: values[index] if index < len(values) else None
                for index in range(len(headers))
                if headers[index]
            }
            normalized_row = self._normalize_tabular_row(raw_row)
            if normalized_row is not None:
                rows.append(normalized_row)

        return rows

    def _extract_rows_for_selected_table(self, payload, config):
        if isinstance(payload, list):
            return payload

        if not isinstance(payload, dict):
            raise ValidationError(
                {
                    "detail": (
                        "El contenido importado debe ser una lista de registros "
                        "o un objeto JSON compatible."
                    )
                }
            )

        if "rows" in payload:
            rows = payload.get("rows")
        elif "tables" in payload:
            rows = payload["tables"].get(config["key"])
        else:
            rows = payload.get(config["key"])

        if rows is None:
            raise ValidationError(
                {
                    "detail": (
                        f"No se encontraron registros para la tabla {config['label']} "
                        "en el archivo importado."
                    ),
                    "table": config["key"],
                    "label": config["label"],
                }
            )

        if not isinstance(rows, list):
            raise ValidationError(
                {
                    "detail": (
                        f"La tabla {config['label']} debe contener una lista de registros."
                    ),
                    "table": config["key"],
                    "label": config["label"],
                }
            )

        return rows

    def _extract_tables(self, payload):
        if not isinstance(payload, dict):
            raise ValidationError(
                {"detail": "El contenido importado debe ser un objeto JSON."}
            )

        tables_payload = payload.get("tables", payload)
        if not isinstance(tables_payload, dict):
            raise ValidationError(
                {"detail": "La propiedad 'tables' debe ser un objeto JSON."}
            )

        selected_tables = {
            config["key"]: tables_payload.get(config["key"], [])
            for config in TABLE_CONFIGS
            if config["key"] in tables_payload
        }

        if not selected_tables:
            raise ValidationError(
                {
                    "detail": (
                        "No se encontraron tablas compatibles en el archivo. "
                        "Usa el formato exportado por esta vista."
                    )
                }
            )

        return selected_tables

    def _import_tables(self, tables_payload):
        summary = {
            "processed": 0,
            "created": 0,
            "updated": 0,
            "unchanged": 0,
            "tables": [],
        }
        models_to_reset = []

        with transaction.atomic():
            for config in TABLE_CONFIGS:
                rows = tables_payload.get(config["key"])
                if rows is None:
                    continue

                if not isinstance(rows, list):
                    raise ValidationError(
                        {
                            "detail": (
                                f"La tabla {config['label']} debe ser una lista de registros."
                            ),
                            "table": config["key"],
                            "label": config["label"],
                        }
                    )

                table_summary = {
                    "key": config["key"],
                    "label": config["label"],
                    "processed": 0,
                    "created": 0,
                    "updated": 0,
                    "unchanged": 0,
                }
                seen_ids = set()

                for row_index, row in enumerate(rows, start=1):
                    cleaned = self._clean_row(config, row, row_index, seen_ids)
                    record_id = cleaned["id"]
                    instance = config["model"].objects.filter(pk=record_id).first()

                    if instance is None:
                        instance = config["model"](id=record_id)
                        self._apply_changes(config, instance, cleaned)
                        self._save_instance(
                            config,
                            instance,
                            row_index=row_index,
                            record_id=record_id,
                            force_insert=True,
                        )
                        models_to_reset.append(config["model"])
                        table_summary["created"] += 1
                    else:
                        changed = self._apply_changes(config, instance, cleaned)
                        if changed:
                            self._save_instance(
                                config,
                                instance,
                                row_index=row_index,
                                record_id=record_id,
                            )
                            table_summary["updated"] += 1
                        else:
                            table_summary["unchanged"] += 1

                    table_summary["processed"] += 1

                summary["processed"] += table_summary["processed"]
                summary["created"] += table_summary["created"]
                summary["updated"] += table_summary["updated"]
                summary["unchanged"] += table_summary["unchanged"]
                summary["tables"].append(table_summary)

            if models_to_reset:
                self._reset_sequences(models_to_reset)

        return summary

    def _clean_row(self, config, row, row_index, seen_ids):
        if not isinstance(row, dict):
            raise ValidationError(
                {
                    "detail": "Cada registro importado debe ser un objeto JSON.",
                    "table": config["key"],
                    "label": config["label"],
                    "row": row_index,
                }
            )

        raw_id = row.get("id")
        if raw_id in (None, ""):
            raise ValidationError(
                {
                    "detail": (
                        "Cada registro debe incluir una clave primaria en el campo 'id'."
                    ),
                    "table": config["key"],
                    "label": config["label"],
                    "row": row_index,
                }
            )

        record_id = self._clean_field_value(
            config,
            field_name="id",
            raw_value=raw_id,
            row_index=row_index,
            record_id=raw_id,
        )

        if record_id in seen_ids:
            raise ValidationError(
                {
                    "detail": (
                        f"El id {record_id} esta duplicado dentro de la tabla importada."
                    ),
                    "table": config["key"],
                    "label": config["label"],
                    "row": row_index,
                    "id": record_id,
                }
            )
        seen_ids.add(record_id)

        cleaned = {"id": record_id}
        for field_name in config["fields"]:
            if field_name == "id" or field_name not in row:
                continue

            if field_name in config.get("foreign_keys", {}):
                cleaned[field_name] = self._resolve_related_instance(
                    config=config,
                    field_name=field_name,
                    raw_value=row[field_name],
                    row_index=row_index,
                    record_id=record_id,
                )
                continue

            cleaned[field_name] = self._clean_field_value(
                config=config,
                field_name=field_name,
                raw_value=row[field_name],
                row_index=row_index,
                record_id=record_id,
            )

        return cleaned

    def _clean_field_value(self, config, field_name, raw_value, row_index, record_id):
        model_field = config["model"]._meta.get_field(field_name)
        try:
            return model_field.clean(raw_value, None)
        except DjangoValidationError as exc:
            raise ValidationError(
                {
                    "detail": "Uno de los campos del registro importado no es valido.",
                    "table": config["key"],
                    "label": config["label"],
                    "row": row_index,
                    "id": record_id,
                    "field": field_name,
                    "errors": exc.messages,
                }
            ) from exc

    def _resolve_related_instance(
        self,
        config,
        field_name,
        raw_value,
        row_index,
        record_id,
    ):
        if raw_value in (None, ""):
            return None

        related_model = config["foreign_keys"][field_name]
        related_field = related_model._meta.pk

        try:
            related_id = related_field.clean(raw_value, None)
        except DjangoValidationError as exc:
            raise ValidationError(
                {
                    "detail": "La clave foranea enviada no es valida.",
                    "table": config["key"],
                    "label": config["label"],
                    "row": row_index,
                    "id": record_id,
                    "field": field_name,
                    "errors": exc.messages,
                }
            ) from exc

        related_instance = related_model.objects.filter(pk=related_id).first()
        if related_instance is None:
            raise ValidationError(
                {
                    "detail": (
                        f"No existe el registro relacionado {related_model.__name__} "
                        f"con id {related_id}."
                    ),
                    "table": config["key"],
                    "label": config["label"],
                    "row": row_index,
                    "id": record_id,
                    "field": field_name,
                }
            )

        return related_instance

    def _apply_changes(self, config, instance, cleaned):
        changed = instance.pk is None

        for field_name, value in cleaned.items():
            if field_name == "id":
                continue

            if field_name in config.get("foreign_keys", {}):
                current_value = getattr(instance, f"{field_name}_id", None)
                incoming_value = value.pk if value else None
            else:
                current_value = getattr(instance, field_name)
                incoming_value = value

            if current_value != incoming_value:
                changed = True

            setattr(instance, field_name, value)

        return changed

    def _save_instance(
        self,
        config,
        instance,
        row_index,
        record_id,
        force_insert=False,
    ):
        try:
            instance.full_clean()
            instance.save(force_insert=force_insert)
        except DjangoValidationError as exc:
            raise ValidationError(
                {
                    "detail": "No se pudo validar el registro importado.",
                    "table": config["key"],
                    "label": config["label"],
                    "row": row_index,
                    "id": record_id,
                    "errors": getattr(exc, "message_dict", exc.messages),
                }
            ) from exc
        except IntegrityError as exc:
            raise ValidationError(
                {
                    "detail": "No se pudo guardar el registro importado.",
                    "table": config["key"],
                    "label": config["label"],
                    "row": row_index,
                    "id": record_id,
                    "errors": [str(exc)],
                }
            ) from exc

    def _reset_sequences(self, models_to_reset):
        unique_models = []
        seen = set()

        for model in models_to_reset:
            if model in seen:
                continue
            seen.add(model)
            unique_models.append(model)

        sql_statements = connection.ops.sequence_reset_sql(no_style(), unique_models)
        if not sql_statements:
            return

        with connection.cursor() as cursor:
            for sql in sql_statements:
                cursor.execute(sql)

    def _get_table_config(self, table_key):
        config = TABLE_CONFIG_MAP.get(table_key)
        if config is None:
            raise ValidationError(
                {
                    "detail": (
                        f"La tabla '{table_key}' no es compatible con esta vista."
                    )
                }
            )
        return config

    @staticmethod
    def _detect_uploaded_format(filename):
        normalized = (filename or "").lower()
        if normalized.endswith(".csv"):
            return "csv"
        if normalized.endswith(".xlsx"):
            return "xlsx"
        if normalized.endswith(".json"):
            return "json"
        raise ParseError("Formato no compatible. Usa archivos CSV, XLSX o JSON.")

    @staticmethod
    def _normalize_tabular_row(row):
        normalized = {}
        has_values = False

        for key, value in row.items():
            if key is None:
                continue

            normalized_key = str(key).strip()
            if not normalized_key:
                continue

            normalized[normalized_key] = value
            if value not in (None, ""):
                has_values = True

        if not has_values:
            return None

        return normalized

    @staticmethod
    def _serialize_tabular_value(value):
        if value is None:
            return ""
        if hasattr(value, "isoformat"):
            return value.isoformat()
        return str(value) if not isinstance(value, bool) else value

    @staticmethod
    def _sheet_title(label):
        return label[:31] or "Datos"

    @staticmethod
    def _is_truthy(value):
        return str(value).lower() in {"1", "true", "yes", "si"}
